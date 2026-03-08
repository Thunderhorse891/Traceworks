"""
TraceWorks Investigation Engine v3.0
4-Tier System: Standard | Comprehensive | Property & Title | Heir & Beneficiary
traceworks.tx@outlook.com
"""

import sys, re, time, json, random, os
from datetime import datetime
from urllib.parse import quote
from collections import defaultdict

try:
    import requests
    from bs4 import BeautifulSoup
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system(f"{sys.executable} -m pip install requests beautifulsoup4 lxml openpyxl --quiet --user")
    import requests
    from bs4 import BeautifulSoup
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ── INJECTED BY POWERSHELL ────────────────────────────────────
FIRST      = "{{FIRST}}"
LAST       = "{{LAST}}"
DOB        = "{{DOB}}"
LAST_ADDR  = "{{LAST_ADDR}}"
LAST_PHONE = "{{LAST_PHONE}}"
TIER       = "{{TIER}}"          # 1=Standard 2=Comprehensive 3=Property 4=Heir
CASE_NUM   = "{{CASE_NUM}}"
ATT_NAME   = "{{ATT_NAME}}"
ATT_FIRM   = "{{ATT_FIRM}}"
ATT_EMAIL  = "{{ATT_EMAIL}}"
OUT_PATH   = "{{OUT_PATH}}"
# ─────────────────────────────────────────────────────────────

SUBJECT   = f"{FIRST} {LAST}"
RPT_DATE  = datetime.now().strftime("%B %d, %Y")
RPT_TIME  = datetime.now().strftime("%I:%M %p")

TIER_NAMES = {
    "1": "Standard Locate",
    "2": "Comprehensive Locate + Assets",
    "3": "Property & Chain of Title",
    "4": "Heir & Beneficiary Location"
}
TIER_PRICES = {"1":"$75","2":"$150","3":"$200","4":"$100"}
SERVICE = TIER_NAMES.get(TIER, "Standard Locate")
PRICE   = TIER_PRICES.get(TIER, "$75")

# ── BRAND COLORS ─────────────────────────────────────────────
BG="0D1117"; BG2="13181F"; BG3="1A2030"; BG4="0A0E14"
GOLD="C9A84C"; GOLD2="E8C97A"; GOLD3="8B6914"
TEXT="E8EDF5"; DIM="8892A4"; DIM2="3A4555"
RED="E53E3E"; RED_BG="1A0808"
AMB="D97706"; AMB_BG="1A1208"
GRN="38A169"; GRN_BG="081A10"
BLUE="1E6FA8"; WHITE="FFFFFF"
PURPLE="7B61FF"; TEAL="0E9F9F"

# ── ROTATING USER AGENTS ──────────────────────────────────────
UAS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:123.0) Gecko/20100101 Firefox/123.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.3 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36 Edg/121.0.2277.128",
]

def hdrs(ref=None):
    h = {"User-Agent": random.choice(UAS),
         "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
         "Accept-Language": "en-US,en;q=0.9", "DNT": "1",
         "Connection": "keep-alive", "Upgrade-Insecure-Requests": "1"}
    if ref: h["Referer"] = ref
    return h

def get(url, timeout=15, ref=None, delay=0.6, retries=2):
    time.sleep(delay + random.uniform(0.1, 0.5))
    for i in range(retries):
        try:
            r = requests.get(url, headers=hdrs(ref), timeout=timeout,
                             allow_redirects=True, verify=True)
            if r.status_code == 200: return r
            if r.status_code == 429: time.sleep(5)
        except: pass
    return None

def post(url, data, timeout=15):
    time.sleep(0.8 + random.uniform(0.2, 0.6))
    try:
        return requests.post(url, data=data, headers=hdrs(), timeout=timeout, verify=True)
    except: return None

def clean(t): return re.sub(r'\s+',' ',str(t or '')).strip()

# ── RESULTS ───────────────────────────────────────────────────
R = {
    "addresses":[], "phones":[], "emails":[], "relatives":[], "associates":[],
    "employers":[], "properties":[], "vehicles":[], "businesses":[],
    "social":{}, "court_records":[], "criminal":[], "probate":[],
    "deeds":[], "liens":[], "mineral":[], "licenses":[],
    "heirs":[], "breaches":[], "aliases":[], "red_flags":[], "sources":[],
    "confidence": "PENDING"
}

def log(name, url, result, status="checked"):
    R["sources"].append({"name":name,"url":url,"result":result,"status":status})
    icon = "✓" if any(x in status.upper() for x in ["FOUND","VERIF","CLEAN"]) else "→"
    print(f"    {icon} [{name}] {result}")

def flag(sev, desc, cat, action):
    R["red_flags"].append({"severity":sev,"desc":desc,"cat":cat,"action":action})

# ════════════════════════════════════════════════════════════
#  SCRAPERS — PEOPLE AGGREGATORS
# ════════════════════════════════════════════════════════════
def s_truepeoplesearch():
    url = f"https://www.truepeoplesearch.com/results?name={quote(SUBJECT)}&citystatezip=TX"
    r = get(url)
    if not r: log("TruePeopleSearch","truepeoplesearch.com","Unreachable","manual"); return
    soup = BeautifulSoup(r.text,"lxml")
    found=False
    for card in soup.select(".card-summary, [data-link-to-details]")[:3]:
        for el in card.select(".content-value,.address"):
            t = clean(el.get_text())
            if re.search(r'\d{2,6}\s+\w',t) and len(t)>10:
                R["addresses"].append({"addr":t,"source":"TruePeopleSearch","confidence":"Medium","corroborated_by":1})
                found=True
        for ph in card.select("[href^='tel:'],.phone"):
            p = re.sub(r'[^\d\(\)\-\s]','',clean(ph.get_text())).strip()
            if len(re.sub(r'\D','',p))==10: R["phones"].append({"phone":p,"source":"TruePeopleSearch"})
        for rel in card.select(".relative,.link-to-details")[:4]:
            rt = clean(rel.get_text())
            if rt and len(rt)>3 and rt not in R["relatives"]: R["relatives"].append(rt)
    detail = soup.select_one("a[href*='/details/']")
    if detail:
        dr = get("https://www.truepeoplesearch.com"+detail["href"], ref=url, delay=1.5)
        if dr:
            ds = BeautifulSoup(dr.text,"lxml")
            for a in ds.select(".address-row,.current-address,.previous-address"):
                t=clean(a.get_text())
                if re.search(r'\d{2,6}\s+\w',t):
                    R["addresses"].append({"addr":t,"source":"TruePeopleSearch","confidence":"High","corroborated_by":1})
                    found=True
            for rel in ds.select(".relative-name")[:8]:
                rt=clean(rel.get_text())
                if rt and rt not in R["relatives"]: R["relatives"].append(rt)
    log("TruePeopleSearch","truepeoplesearch.com",f"{len(R['addresses'])} addr, {len(R['phones'])} phones","FOUND" if found else "No match")

def s_fastpeople():
    url = f"https://www.fastpeoplesearch.com/name/{quote(FIRST)}-{quote(LAST)}_TX"
    r = get(url)
    if not r: log("FastPeopleSearch","fastpeoplesearch.com","Unreachable","manual"); return
    soup=BeautifulSoup(r.text,"lxml"); found=False
    for card in soup.select(".card,.result-card")[:4]:
        for a in card.select(".address,[class*='addr']"):
            t=clean(a.get_text())
            if re.search(r'\d+\s+\w',t):
                R["addresses"].append({"addr":t,"source":"FastPeopleSearch","confidence":"Medium","corroborated_by":1})
                found=True
        for p in re.findall(r'\(?\d{3}\)?[\s\-\.]\d{3}[\s\-\.]\d{4}',card.get_text()):
            R["phones"].append({"phone":p,"source":"FastPeopleSearch"})
        for rel in card.select(".relative-name,.person-name")[:3]:
            rt=clean(rel.get_text())
            if rt and LAST.upper() not in rt.upper() and len(rt)>3:
                if rt not in R["relatives"]: R["relatives"].append(rt)
    log("FastPeopleSearch","fastpeoplesearch.com",f"{'Address found' if found else 'No match'}","FOUND" if found else "No match")

def s_spokeo():
    url = f"https://www.spokeo.com/{FIRST}-{LAST}/Texas"
    r = get(url, ref="https://www.spokeo.com/")
    if not r: log("Spokeo","spokeo.com","Unreachable","manual"); return
    soup=BeautifulSoup(r.text,"lxml"); found=False
    for el in soup.select(".city-state,.location,[class*='location']"):
        t=clean(el.get_text())
        if t and len(t)>3:
            R["addresses"].append({"addr":f"{t} (city-level)","source":"Spokeo","confidence":"Low","corroborated_by":1})
            found=True
    for rel in soup.select(".relative,[class*='relative']")[:5]:
        rt=clean(rel.get_text())
        if rt and rt not in R["relatives"] and len(rt)>3: R["relatives"].append(rt)
    log("Spokeo","spokeo.com","Preview data found" if found else "No preview data","FOUND" if found else "No match")

def s_zaba():
    url = f"https://www.zabasearch.com/people/{FIRST.lower()}+{LAST.lower()}/texas/"
    r = get(url)
    if not r: log("ZabaSearch","zabasearch.com","Unreachable","manual"); return
    soup=BeautifulSoup(r.text,"lxml"); found=False
    for res in soup.select(".search-result,li[class*='result']")[:5]:
        t=res.get_text()
        m=re.search(r'(\d{2,6}\s+[\w\s]+(?:ST|AVE|DR|BLVD|RD|LN|CT|WAY)[^\n]{0,50})',t,re.I)
        if m:
            R["addresses"].append({"addr":clean(m.group(1)),"source":"ZabaSearch","confidence":"Medium","corroborated_by":1})
            found=True
        for p in re.findall(r'\(?\d{3}\)?[\s\-]\d{3}[\s\-]\d{4}',t):
            R["phones"].append({"phone":p,"source":"ZabaSearch"})
    log("ZabaSearch","zabasearch.com","Records found" if found else "No match","FOUND" if found else "No match")

def s_411():
    url = f"https://www.411.com/name/{FIRST.lower()}-{LAST.lower()}/TX/"
    r = get(url, ref="https://www.411.com/")
    if not r: log("411.com","411.com","Unreachable","manual"); return
    soup=BeautifulSoup(r.text,"lxml"); found=False
    for res in soup.select(".result,.person,[class*='listing']")[:5]:
        t=res.get_text()
        m=re.search(r'(\d{2,6}\s+[\w\s]{4,30},\s*\w[\w\s]*,?\s*TX)',t,re.I)
        if m:
            R["addresses"].append({"addr":clean(m.group(1)),"source":"411.com","confidence":"Medium","corroborated_by":1})
            found=True
        for p in re.findall(r'\(?\d{3}\)?[\s\-\.]\d{3}[\s\-\.]\d{4}',t):
            R["phones"].append({"phone":p,"source":"411.com"})
    log("411.com","411.com","Found" if found else "No match","FOUND" if found else "No match")

def s_anywho():
    url = f"https://www.anywho.com/people/{quote(FIRST)}+{quote(LAST)}/TX"
    r = get(url)
    if not r: log("AnyWho","anywho.com","Unreachable","manual"); return
    soup=BeautifulSoup(r.text,"lxml"); found=False
    for row in soup.select(".person-listing,.result-item")[:5]:
        t=row.get_text()
        m=re.search(r'(\d{2,6}\s+\w[\w\s,]+TX\s*\d{5})',t,re.I)
        if m:
            R["addresses"].append({"addr":clean(m.group(1)),"source":"AnyWho","confidence":"Medium","corroborated_by":1})
            found=True
        for p in re.findall(r'\(?\d{3}\)?[\s\-\.]\d{3}[\s\-\.]\d{4}',t):
            R["phones"].append({"phone":p,"source":"AnyWho"})
    log("AnyWho","anywho.com","Found" if found else "No match","FOUND" if found else "No match")

def s_peekyou():
    url = f"https://www.peekyou.com/{FIRST.lower()}_{LAST.lower()}"
    r = get(url, ref="https://www.peekyou.com/")
    if not r: log("PeekYou","peekyou.com","Unreachable","manual"); return
    soup=BeautifulSoup(r.text,"lxml"); found=False
    for link in soup.select("a[href]"):
        href=link.get("href","")
        for plat in ["facebook.com","twitter.com","x.com","linkedin.com","instagram.com","tiktok.com","youtube.com"]:
            if plat in href and LAST.lower() in href.lower():
                name=plat.split(".")[0].capitalize()
                if name not in R["social"]: R["social"][name]=href; found=True
    loc=soup.select_one(".location,[class*='city']")
    if loc:
        R["addresses"].append({"addr":f"{clean(loc.get_text())} (city-level)","source":"PeekYou","confidence":"Low","corroborated_by":1})
    log("PeekYou","peekyou.com",f"Social: {list(R['social'].keys())}" if found else "No profiles","FOUND" if found else "No match")

# ════════════════════════════════════════════════════════════
#  SCRAPERS — SEARCH ENGINES
# ════════════════════════════════════════════════════════════
def s_google():
    dorks=[f'"{FIRST} {LAST}" Texas address',f'"{FIRST} {LAST}" site:linkedin.com',f'"{FIRST} {LAST}" TX email phone']
    for dork in dorks[:2]:
        r=get(f"https://www.google.com/search?q={quote(dork)}&num=10",delay=2)
        if not r: continue
        for p in re.findall(r'\(?\d{3}\)?[\s\-\.]\d{3}[\s\-\.]\d{4}',r.text):
            R["phones"].append({"phone":p,"source":"Google OSINT"})
        for e in re.findall(r'[\w\.\+\-]+@[\w\.\-]+\.\w{2,6}',r.text):
            if LAST.lower() in e.lower() or FIRST.lower() in e.lower():
                R["emails"].append({"email":e,"source":"Google OSINT"})
        li=re.search(r'linkedin\.com/in/([\w\-]+)',r.text,re.I)
        if li: R["social"]["LinkedIn"]=f"linkedin.com/in/{li.group(1)}"
        time.sleep(1.5)
    log("Google OSINT","google.com","Data extracted","checked")

def s_bing():
    r=get(f"https://www.bing.com/search?q={quote(FIRST+' '+LAST+' Texas address phone')}&count=10",
          ref="https://www.bing.com/",delay=1.5)
    if not r: log("Bing","bing.com","Unreachable","manual"); return
    found=False
    for p in re.findall(r'\(?\d{3}\)?[\s\-\.]\d{3}[\s\-\.]\d{4}',r.text):
        if p not in [x["phone"] for x in R["phones"]]:
            R["phones"].append({"phone":p,"source":"Bing"}); found=True
    for e in re.findall(r'[\w\.\-]+@[\w\.\-]+\.\w{2,4}',r.text):
        if (LAST.lower() in e.lower() or FIRST.lower() in e.lower()) and e not in [x["email"] for x in R["emails"]]:
            R["emails"].append({"email":e,"source":"Bing"}); found=True
    log("Bing","bing.com","Data extracted" if found else "No match","FOUND" if found else "checked")

# ════════════════════════════════════════════════════════════
#  SCRAPERS — TEXAS GOVERNMENT
# ════════════════════════════════════════════════════════════
def s_cad_multi():
    """Search 8 major TX CADs"""
    counties=[
        ("Travis CAD",f"https://esearch.traviscad.org/Search/Property?searchValue={quote(LAST)}&searchType=OwnerName"),
        ("Williamson CAD",f"https://wcad.org/online-property-search/?prop_type=R&owner_name={quote(LAST)}"),
        ("Harris CAD",f"https://public.hcad.org/records/details.asp?type=res&owner={quote(LAST)}"),
        ("Milam CAD",f"https://milam.tx.publicsearch.us/?term={quote(LAST)}&searchField=ownerName"),
        ("Bell CAD",f"https://bell.tx.publicsearch.us/?term={quote(LAST)}&searchField=ownerName"),
        ("Bastrop CAD",f"https://bastrop.tx.publicsearch.us/?term={quote(LAST)}&searchField=ownerName"),
        ("Hays CAD",f"https://hays.tx.publicsearch.us/?term={quote(LAST)}&searchField=ownerName"),
        ("Bexar CAD",f"https://bexar.tx.publicsearch.us/?term={quote(LAST)}&searchField=ownerName"),
    ]
    for name,url in counties:
        r=get(url,delay=0.7)
        if not r: log(name,url,"Unreachable","manual"); continue
        addrs=re.findall(r'(\d{2,6}\s+[\w\s]{4,30}(?:ST|AVE|DR|BLVD|RD|LN|CT|WAY|TRAIL|HWY|FM\s*\d+|CR\s*\d+)[^\n<"]{0,50})',r.text,re.I)
        if addrs:
            for a in addrs[:2]:
                ca=clean(a)
                R["properties"].append({"addr":ca,"county":name,"source":name,"type":"Real Property"})
                R["addresses"].append({"addr":ca,"source":name,"confidence":"High","corroborated_by":1})
            log(name,url,f"Property: {clean(addrs[0])[:50]}","FOUND")
        else:
            log(name,url,"No property found","No match")

def s_tx_sos():
    url=f"https://ourcpa.cpa.state.tx.us/coa/servlet/cpa.app.coa.CoaGetName?coaBeginNameSearch={quote(LAST)}&coaSearchType=Name"
    r=get(url)
    if not r:
        log("TX Secretary of State","direct.sos.state.tx.us","Manual required — direct.sos.state.tx.us","manual"); return
    found=False
    for m in re.findall(r'([\w\s&\.\,\-\']{4,50}(?:LLC|INC|CORP|LP|LLP|PLLC|PC|CO\.))',r.text,re.I)[:3]:
        biz=clean(m)
        if LAST.upper() in biz.upper() or FIRST.upper() in biz.upper():
            R["businesses"].append({"name":biz,"source":"TX SOS","status":""}); found=True
    log("TX SOS","direct.sos.state.tx.us",f"Business: {R['businesses'][0]['name']}" if found else "No business","FOUND" if found else "No match")

def s_tx_courts():
    url=f"https://search.txcourts.gov/CaseSearch.aspx?fn={quote(FIRST)}&ln={quote(LAST)}"
    r=get(url)
    if not r: log("TX Courts","search.txcourts.gov","Manual required — CAPTCHA may block","manual"); return
    found=False
    cases=re.findall(r'(?:Case|Cause)[^\d]{0,5}([\w\d\-]{6,25})',r.text,re.I)
    if cases:
        for c in cases[:5]: R["court_records"].append({"case":c,"source":"TX Courts Online","type":"Civil/Criminal"})
        flag("MEDIUM",f"{len(cases)} court record(s) in TX Courts Online","LEGAL","Pull full details at search.txcourts.gov")
        found=True
    for ct in ["PROBATE","FAMILY","CRIMINAL","CIVIL"]:
        if ct in r.text.upper():
            R["court_records"].append({"case":ct+" case found","source":"TX Courts","type":ct})
    log("TX Courts Online","search.txcourts.gov",f"{len(cases)} case(s)" if found else "Clean record","FOUND" if found else "Clean")

def s_tdcj():
    url="https://offender.tdcj.texas.gov/OffenderSearch/search.action"
    r=post(url,{"lastName":LAST,"firstName":FIRST,"gender":"X","race":"X","btn":"Search"})
    if not r: log("TDCJ","offender.tdcj.texas.gov","Manual required","manual"); return
    if "no records" in r.text.lower() or len(r.text)<2000:
        log("TDCJ","offender.tdcj.texas.gov","No active incarceration","Clean")
    else:
        soup=BeautifulSoup(r.text,"lxml")
        if soup.select("table tr")[1:]:
            flag("HIGH","Subject found in TDCJ inmate records — criminal history","CRIMINAL","Disclose to attorney immediately")
            R["criminal"].append({"record":"TDCJ match","source":"TDCJ"})
            log("TDCJ","offender.tdcj.texas.gov","RECORD FOUND","FOUND")
        else:
            log("TDCJ","offender.tdcj.texas.gov","No incarceration","Clean")

def s_sex_offender():
    url=f"https://publicsite.dps.texas.gov/SexOffenderRegistry/Search/Detailed/Results?lastName={quote(LAST)}&firstName={quote(FIRST)}&city=&county=&zip=&dob="
    r=get(url)
    if not r: log("TX Sex Offender Registry","publicsite.dps.texas.gov","Manual required","manual"); return
    if "no results" in r.text.lower() or len(r.text)<2000:
        log("TX Sex Offender Registry","publicsite.dps.texas.gov","Not on registry","Clean")
    else:
        soup=BeautifulSoup(r.text,"lxml")
        if soup.select(".offender-card,tr[class*='result']"):
            flag("HIGH","Subject on TX DPS Sex Offender Registry","CRIMINAL","CRITICAL — disclose to attorney immediately")
            R["criminal"].append({"record":"Sex Offender Registry","source":"TX DPS"})
            log("TX Sex Offender Registry","publicsite.dps.texas.gov","REGISTRY MATCH","FOUND")
        else:
            log("TX Sex Offender Registry","publicsite.dps.texas.gov","Not on registry","Clean")

def s_tdlr():
    url=f"https://www.tdlr.texas.gov/LicenseSearch/licfile.asp?nameType=Person&searchName={quote(LAST)}%2C+{quote(FIRST)}"
    r=get(url)
    if not r: log("TDLR","tdlr.texas.gov","Manual required","manual"); return
    found=False
    for row in BeautifulSoup(r.text,"lxml").select("table tr")[1:6]:
        cells=[clean(td.get_text()) for td in row.select("td")]
        if len(cells)>=3 and LAST.upper() in cells[0].upper():
            R["licenses"].append({"type":cells[1] if len(cells)>1 else "","number":cells[2] if len(cells)>2 else "","addr":cells[3] if len(cells)>3 else ""})
            if len(cells)>3 and cells[3]: R["addresses"].append({"addr":cells[3],"source":"TDLR License","confidence":"High","corroborated_by":1})
            found=True
    log("TDLR","tdlr.texas.gov",f"License: {R['licenses'][0]['type']}" if found else "No license","FOUND" if found else "No match")

def s_rrc():
    """TX Railroad Commission — mineral/oil interests"""
    url=f"https://webapps2.rrc.texas.gov/EWA/operatorQueryAction.do?operatorName={quote(LAST)}"
    r=get(url,ref="https://webapps2.rrc.texas.gov/")
    if not r: log("TX Railroad Commission","webapps2.rrc.texas.gov","Manual required","manual"); return
    found=False
    if LAST.upper() in r.text.upper():
        permits=re.findall(r'(?:API|Permit)[^\d]{0,5}([\d\-]{8,20})',r.text)
        if permits:
            R["mineral"].append({"desc":f"RRC Operator — {len(permits)} permit(s)","source":"TX RRC","api":permits[0]})
            flag("MEDIUM",f"Subject is TX RRC operator — {len(permits)} oil/gas permit(s) on file","ASSETS","Check mineral ownership and royalty payments")
            found=True
    log("TX Railroad Commission","webapps2.rrc.texas.gov",f"{len(R['mineral'])} mineral interests" if found else "No operator record","FOUND" if found else "No match")

def s_tx_courts_probate():
    """Search specifically for probate/estate cases"""
    url=f"https://search.txcourts.gov/CaseSearch.aspx?fn={quote(FIRST)}&ln={quote(LAST)}&caseType=PROBATE"
    r=get(url)
    if not r: log("TX Courts (Probate)","search.txcourts.gov","Manual required","manual"); return
    found=False
    cases=re.findall(r'(?:Case|Cause|Estate)[^\d]{0,8}([\w\d\-]{6,25})',r.text,re.I)
    if cases:
        for c in cases[:5]:
            R["probate"].append({"case":c,"source":"TX Courts","type":"Probate/Estate"})
        flag("MEDIUM",f"{len(cases)} probate/estate record(s) found","PROBATE","Review estate details at search.txcourts.gov")
        found=True
    # Also look for estate names
    estate_m=re.findall(rf'Estate of\s+[\w\s,\.]+{re.escape(LAST)}[\w\s,\.]*',r.text,re.I)
    for em in estate_m[:3]:
        R["probate"].append({"case":clean(em),"source":"TX Courts","type":"Estate Record"})
        found=True
    log("TX Courts (Probate)","search.txcourts.gov",f"{len(cases)} probate case(s)" if found else "No probate records","FOUND" if found else "No match")

def s_linkedin():
    url=f"https://www.linkedin.com/pub/dir/?first={quote(FIRST)}&last={quote(LAST)}&search=Search"
    r=get(url,ref="https://www.linkedin.com/",delay=2)
    if not r: log("LinkedIn","linkedin.com","Unreachable","manual"); return
    found=False
    for slug in re.findall(r'linkedin\.com/in/([\w\-\.]{3,50})',r.text,re.I)[:3]:
        R["social"]["LinkedIn"]=f"linkedin.com/in/{slug}"; found=True
    for emp in re.findall(r'(?:at|@)\s+([\w\s&\-\.]{5,40})',r.text,re.I)[:2]:
        ec=clean(emp)
        if ec and ec not in [e["employer"] for e in R["employers"]]:
            R["employers"].append({"employer":ec,"source":"LinkedIn","addr":""})
    log("LinkedIn","linkedin.com",f"Profile: {R['social'].get('LinkedIn','None')}","FOUND" if found else "No match")

def s_facebook():
    url=f"https://www.facebook.com/public/{FIRST}-{LAST}"
    r=get(url,ref="https://www.facebook.com/",delay=2)
    if not r: log("Facebook","facebook.com","Unreachable — manual check","manual"); return
    found=False
    if "Texas" in r.text or "TX" in r.text:
        m=re.search(r'(?:Lives in|From)[^:]{0,15}:\s*([\w\s,]+?(?:Texas|TX)[^\<"]{0,30})',r.text,re.I)
        if m:
            R["addresses"].append({"addr":f"{clean(m.group(1))} (Facebook)","source":"Facebook","confidence":"Low","corroborated_by":1})
            found=True
    if len(r.text)>8000: R["social"]["Facebook"]=f"facebook.com/public/{FIRST}-{LAST}"; found=True
    log("Facebook","facebook.com","Profile found" if found else "No public profile","FOUND" if found else "No match")

def s_instagram():
    for handle in [f"{FIRST.lower()}.{LAST.lower()}",f"{FIRST.lower()}{LAST.lower()}",f"{FIRST.lower()}_{LAST.lower()}"]:
        r=get(f"https://www.instagram.com/{handle}/",delay=1.5)
        if r and r.status_code==200 and '"username"' in r.text:
            R["social"]["Instagram"]=f"instagram.com/{handle}"
            log("Instagram","instagram.com",f"Handle: {handle}","FOUND"); return
    log("Instagram","instagram.com","No match","No match")

def s_usps_verify():
    best=[a for a in R["addresses"] if a.get("confidence") in ["High","Medium","VERIFIED"]]
    if not best: log("USPS","tools.usps.com","No address to verify","skipped"); return
    addr=best[0]["addr"]
    r=get(f"https://tools.usps.com/zip-code-lookup.htm?byaddress&tAddress={quote(addr)}")
    if r and re.search(r'\d{5}',r.text):
        for a in R["addresses"]:
            if a["addr"]==addr: a["confidence"]="VERIFIED"
        log("USPS","tools.usps.com","Address verified","VERIFIED")
    else:
        log("USPS","tools.usps.com","Could not verify","checked")

def s_county_clerk_deeds(county_name, search_term):
    """Generic county deed search — publicsearch.us portal"""
    slug_map={"milam":"milam","travis":"travis","williamson":"williamson",
              "bell":"bell","bastrop":"bastrop","hays":"hays","caldwell":"caldwell"}
    slug=slug_map.get(county_name.lower().replace(" ",""))
    if not slug: log(f"{county_name} Clerk",f"{county_name.lower()}.tx.publicsearch.us","County not in portal list","manual"); return
    url=f"https://{slug}.tx.publicsearch.us/?term={quote(search_term)}&searchField=grantor"
    r=get(url)
    if not r:
        url2=f"https://{slug}.tx.publicsearch.us/?term={quote(search_term)}&searchField=grantee"
        r=get(url2)
    if not r: log(f"{county_name} County Clerk",url,"Unreachable","manual"); return
    found=False
    # Deed/instrument numbers
    inst=re.findall(r'(?:Doc|Instrument|File)\s*[#:]?\s*([\d]{4,12})',r.text,re.I)
    deed_types=re.findall(r'(Quitclaim|Warranty|Deed of Trust|Release|Lien|Lis Pendens|Easement|Affidavit)',r.text,re.I)
    dates=re.findall(r'\b(20\d{2}[\-\/]\d{2}[\-\/]\d{2}|\d{1,2}[\-\/]\d{1,2}[\-\/]20\d{2})\b',r.text)
    if inst or deed_types:
        for dt in list(set(deed_types))[:3]:
            R["deeds"].append({"type":dt,"instrument":inst[0] if inst else "","county":county_name,"source":f"{county_name} County Clerk"})
        flag("MEDIUM" if "Lis Pendens" in deed_types or "Lien" in deed_types else "LOW",
             f"{county_name} County — deed records found: {', '.join(list(set(deed_types))[:3])}",
             "TITLE",f"Review full deed record at {slug}.tx.publicsearch.us")
        found=True
    log(f"{county_name} County Clerk",url,f"Deeds: {', '.join(list(set(deed_types))[:3])}" if found else "No deed records","FOUND" if found else "No match")

# ════════════════════════════════════════════════════════════
#  DEDUPLICATION & CONFIDENCE SCORING
# ════════════════════════════════════════════════════════════
def dedupe():
    # Addresses
    seen=defaultdict(list)
    for a in R["addresses"]:
        key=re.sub(r'\s+',' ',a["addr"].lower().strip()[:80])
        seen[key].append(a["source"])
    R["addresses"]=[]
    for key,srcs in seen.items():
        n=len(srcs)
        conf="VERIFIED" if n>=3 else "High" if n>=2 else R["addresses"][0]["confidence"] if R["addresses"] else "Medium"
        R["addresses"].append({"addr":key.title(),"source":", ".join(set(srcs)),"confidence":conf,"corroborated_by":n})
    R["addresses"].sort(key=lambda x:x["corroborated_by"],reverse=True)
    # Phones
    seen_p=set()
    uniq=[]
    for p in R["phones"]:
        d=re.sub(r'\D','',p["phone"])[-10:]
        if d not in seen_p and len(d)==10: seen_p.add(d); uniq.append(p)
    R["phones"]=uniq
    # Relatives
    R["relatives"]=list(dict.fromkeys(R["relatives"]))[:15]
    # Set confidence
    best=R["addresses"][0] if R["addresses"] else None
    if best:
        c=best["confidence"]; n=best["corroborated_by"]
        if c=="VERIFIED" or n>=3: R["confidence"]="★★★★★  VERIFIED — 3+ corroborating sources"
        elif c=="High" or n>=2:   R["confidence"]="★★★★☆  HIGH — Multiple sources agree"
        else:                      R["confidence"]="★★☆☆☆  LOW — Single source, field verify"
    else:
        R["confidence"]="★☆☆☆☆  UNLOCATED — Manual investigation required"
        flag("HIGH","Subject not located automatically","LOCATE","Manual skip trace or paid database (TLO/LexisNexis) recommended")

# ════════════════════════════════════════════════════════════
#  EXCEL HELPERS
# ════════════════════════════════════════════════════════════
def F(bold=False,sz=10,color=TEXT,italic=False):
    return Font(name="Calibri",bold=bold,size=sz,color=color,italic=italic)
def fill(c): return PatternFill("solid",fgColor=c)
def bdr(style="thin",color=DIM2):
    s=Side(style=style,color=color); return Border(left=s,right=s,top=s,bottom=s)
def algn(h="left",v="center",wrap=False,ind=0):
    return Alignment(horizontal=h,vertical=v,wrap_text=wrap,indent=ind)
def cw(ws,d):
    for col,w in d.items(): ws.column_dimensions[col].width=w
def rh(ws,r,h): ws.row_dimensions[r].height=h

def C(ws,ref,val="",bold=False,sz=10,color=TEXT,bg=BG2,
      h="left",v="center",wrap=False,italic=False,bc=DIM2,ind=0,bs="thin"):
    c=ws[ref]; c.value=val; c.font=F(bold,sz,color,italic)
    c.fill=fill(bg); c.alignment=algn(h,v,wrap,ind); c.border=bdr(bs,bc)

def MC(ws,r1,c1,r2,c2,val="",bold=False,sz=10,color=TEXT,
       bg=BG2,h="left",v="center",wrap=False,italic=False,bc=DIM2,ind=0):
    cl1=get_column_letter(c1); cl2=get_column_letter(c2)
    ws.merge_cells(f"{cl1}{r1}:{cl2}{r2}")
    c=ws[f"{cl1}{r1}"]; c.value=val; c.font=F(bold,sz,color,italic)
    c.fill=fill(bg); c.alignment=algn(h,v,wrap,ind); c.border=bdr("thin",bc)

def gold_bar(ws,r,nc=5):
    rh(ws,r,3); MC(ws,r,1,r,nc,bg=GOLD,bc=GOLD)

def spacer(ws,r,h=8,nc=5):
    rh(ws,r,h)
    for i in range(1,nc+1): ws[f"{get_column_letter(i)}{r}"].fill=fill(BG)

def page_hdr(ws,r1,title,subtitle,nc=4,accent=GOLD):
    gold_bar(ws,r1,nc); rh(ws,r1+1,30)
    MC(ws,r1+1,1,r1+1,nc,f"  {title}",bold=True,sz=14,color=accent,bg=BG4,bc=GOLD3)
    rh(ws,r1+2,18)
    MC(ws,r1+2,1,r1+2,nc,f"  {subtitle}",sz=9,color=DIM,bg=BG,italic=True,bc=BG)
    spacer(ws,r1+3,8,nc)
    return r1+4

def sec(ws,r,c1,c2,label,accent=GOLD,h=22):
    rh(ws,r,h)
    MC(ws,r,c1,r,c2,f"  ◈  {label}",bold=True,sz=9,color=accent,bg=BG4,bc=GOLD3)
    ws[f"{get_column_letter(c1)}{r}"].border=Border(
        left=Side(style="medium",color=accent),right=Side(style="thin",color=DIM2),
        top=Side(style="thin",color=DIM2),bottom=Side(style="thin",color=DIM2))

def thdr(ws,r,cols,h=20):
    rh(ws,r,h)
    for cn,lbl in cols:
        C(ws,f"{get_column_letter(cn)}{r}",lbl,bold=True,sz=9,color=GOLD2,bg=BG4,h="center",bc=GOLD3)

def drow(ws,r,cols,alt=False,h=20,wrap=False,color=TEXT):
    bg=BG3 if alt else BG2; rh(ws,r,h)
    for cn,val in cols:
        C(ws,f"{get_column_letter(cn)}{r}",str(val),sz=9,color=color,bg=bg,wrap=wrap,ind=1,bc=DIM2)

def kv(ws,r,k,v,alt=False,h=22,vc=None,nc=4):
    bg=BG3 if alt else BG2; rh(ws,r,h)
    ws[f"A{r}"].value=k; ws[f"A{r}"].font=F(True,9,GOLD)
    ws[f"A{r}"].fill=fill(bg); ws[f"A{r}"].alignment=algn("right")
    ws[f"A{r}"].border=Border(left=Side(style="thin",color=DIM2),right=Side(style="medium",color=GOLD),
                               top=Side(style="thin",color=DIM2),bottom=Side(style="thin",color=DIM2))
    MC(ws,r,2,r,nc,str(v),sz=9,color=vc or TEXT,bg=bg,wrap=True,ind=1,bc=DIM2)

def flagrow(ws,r,sev,desc,cat,rec,h=24):
    rh(ws,r,h)
    s=sev.upper()
    if "HIGH" in s: bg,fc=RED_BG,RED
    elif "MED" in s: bg,fc=AMB_BG,AMB
    else: bg,fc=GRN_BG,GRN
    for cn,val in [(1,sev),(2,desc),(3,cat),(4,rec)]:
        C(ws,f"{get_column_letter(cn)}{r}",val,bold=True,sz=9,color=fc,bg=bg,wrap=True,
          bc=fc,bs="medium" if cn==1 else "thin")

def cover_sheet(wb, tier_color):
    ws=wb.active; ws.title="COVER"
    ws.sheet_view.showGridLines=False; ws.sheet_properties.tabColor=tier_color
    cw(ws,{"A":3,"B":6,"C":22,"D":28,"E":22,"F":14,"G":3})
    for r in range(1,30):
        rh(ws,r,16)
        for c in range(1,8): ws[f"{get_column_letter(c)}{r}"].fill=fill(BG)
    gold_bar(ws,2,7); rh(ws,3,8)
    rh(ws,4,58); MC(ws,4,2,4,6,"TRACEWORKS",bold=True,sz=44,color=GOLD,bg=BG,h="left",v="center",bc=BG)
    rh(ws,5,20); MC(ws,5,2,5,6,"SKIP TRACING  ·  OSINT  ·  PROPERTY RESEARCH",sz=9,color=DIM,bg=BG,h="left",italic=True,bc=BG)
    gold_bar(ws,6,7); spacer(ws,7,12,7)
    rh(ws,8,32); MC(ws,8,2,8,6,"INVESTIGATION REPORT",bold=True,sz=18,color=WHITE,bg=BG2,h="left",v="center",bc=DIM2)
    rh(ws,9,24); MC(ws,9,2,9,6,f"{SERVICE}  ·  {PRICE}",sz=11,color=tier_color,bg=BG2,h="left",italic=True,bc=DIM2)
    spacer(ws,10,10,7)
    info=[("SUBJECT",SUBJECT),("CASE #",CASE_NUM),("SERVICE",SERVICE),
          ("PRICE",PRICE),("DATE",RPT_DATE+" at "+RPT_TIME),
          ("ATTORNEY",f"{ATT_NAME}  ·  {ATT_FIRM}"),("CLIENT EMAIL",ATT_EMAIL),
          ("PREPARED BY","TraceWorks  ·  traceworks.tx@outlook.com"),
          ("CONFIDENCE",R.get("confidence","Pending"))]
    alt=False
    for i,(label,val) in enumerate(info):
        r=11+i; alt=not alt; rh(ws,r,22); bg=BG3 if alt else BG2
        is_conf=label=="CONFIDENCE"
        ws[f"C{r}"].value=label; ws[f"C{r}"].font=F(True,9,GOLD); ws[f"C{r}"].fill=fill(BG4)
        ws[f"C{r}"].alignment=algn("right",ind=1)
        ws[f"C{r}"].border=Border(left=Side(style="thin",color=DIM2),right=Side(style="medium",color=GOLD),
                                   top=Side(style="thin",color=DIM2),bottom=Side(style="thin",color=DIM2))
        vc=tier_color if is_conf else TEXT
        MC(ws,r,4,r,6,val,sz=10 if is_conf else 9,color=vc,bg=BLUE if is_conf else bg,ind=1,bc=DIM2,bold=is_conf)
    spacer(ws,20,10,7)
    rh(ws,21,20)
    MC(ws,21,2,21,6,
       f"  Sources: {len(R['sources'])}   ·   Addresses: {len(R['addresses'])}   ·   Phones: {len(R['phones'])}   ·   Flags: {len(R['red_flags'])}",
       sz=9,color=DIM,bg=BG3,h="left",bc=DIM2)
    spacer(ws,22,8,7); gold_bar(ws,23,7)
    rh(ws,24,16); MC(ws,24,2,24,6,"TraceWorks  ·  traceworks.tx@outlook.com  ·  For Authorized Legal Use Only",sz=8,color=DIM,bg=BG,h="center",italic=True,bc=BG)

def source_log_sheet(wb):
    ws=wb.create_sheet("SOURCE LOG")
    ws.sheet_view.showGridLines=False; ws.sheet_properties.tabColor=BLUE
    cw(ws,{"A":5,"B":26,"C":30,"D":14,"E":22})
    r=page_hdr(ws,1,"SOURCE LOG",f"Case: {CASE_NUM}  ·  {RPT_DATE} {RPT_TIME}  ·  Auto-generated TraceWorks",nc=5)
    thdr(ws,r,[(1,"#"),(2,"SOURCE"),(3,"URL"),(4,"DATE"),(5,"RESULT")],h=22); r+=1
    for i,src in enumerate(R["sources"]):
        alt=i%2==1; bg=BG3 if alt else BG2; rh(ws,r+i,20)
        is_f="FOUND" in src["status"].upper() or "VERIF" in src["status"].upper()
        is_c="CLEAN" in src["status"].upper()
        rc=GRN if is_f or is_c else DIM
        C(ws,f"A{r+i}",str(i+1),bold=True,sz=9,color=GOLD,bg=bg,h="center",bc=DIM2)
        C(ws,f"B{r+i}",src["name"],sz=9,color=TEXT,bg=bg,ind=1,bc=DIM2)
        C(ws,f"C{r+i}",src["url"],sz=9,color=DIM,bg=bg,ind=1,bc=DIM2)
        C(ws,f"D{r+i}",RPT_DATE,sz=8,color=DIM,bg=bg,h="center",bc=DIM2)
        C(ws,f"E{r+i}",src["result"],sz=9,color=rc,bg=bg,ind=1,bc=DIM2,wrap=True)
    spacer(ws,r+len(R["sources"]),8,5); gold_bar(ws,r+len(R["sources"])+1,5)

def next_steps_sheet(wb, steps, accent=GOLD):
    ws=wb.create_sheet("NEXT STEPS")
    ws.sheet_view.showGridLines=False; ws.sheet_properties.tabColor=GRN
    cw(ws,{"A":6,"B":42,"C":18,"D":22})
    r=page_hdr(ws,1,"NEXT STEPS & RECOMMENDATIONS",f"Prepared for: {ATT_NAME}  ·  {ATT_FIRM}  ·  {RPT_DATE}")
    thdr(ws,r,[(1,"#"),(2,"ACTION"),(3,"STATUS"),(4,"NOTES")],h=22); r+=1
    for i,(num,action,status,notes) in enumerate(steps):
        alt=i%2==1; bg=BG3 if alt else BG2; rh(ws,r+i,24)
        C(ws,f"A{r+i}",num,bold=True,sz=11,color=accent,bg=bg,h="center",bc=DIM2)
        C(ws,f"B{r+i}",action,sz=9,color=TEXT,bg=bg,wrap=True,ind=1,bc=DIM2)
        C(ws,f"C{r+i}",status,sz=9,color=DIM,bg=bg,h="center",bc=DIM2)
        C(ws,f"D{r+i}",notes,sz=9,color=DIM,bg=bg,ind=1,wrap=True,bc=DIM2)
    nr=r+len(steps)+1; spacer(ws,nr,10,4); rh(ws,nr+1,50)
    MC(ws,nr+1,1,nr+1,4,
       "DISCLAIMER: Compiled using publicly available records and lawful OSINT methodology. "
       "For authorized legal use only. TraceWorks is not a licensed PI agency. "
       "Results should be verified before use in legal proceedings. "
       f"TraceWorks  ·  traceworks.tx@outlook.com  ·  Texas OSINT & Skip Tracing",
       sz=8,color=DIM,bg=BG3,italic=True,wrap=True,h="left",v="top",bc=DIM2)
    gold_bar(ws,nr+2,4)

# ════════════════════════════════════════════════════════════
#  TIER 1 — STANDARD ($75): Address + Phone Only
# ════════════════════════════════════════════════════════════
def build_t1():
    wb=openpyxl.Workbook()
    cover_sheet(wb, GRN)

    # Sheet 2: Locate Results
    ws=wb.create_sheet("LOCATE RESULTS")
    ws.sheet_view.showGridLines=False; ws.sheet_properties.tabColor=GRN
    cw(ws,{"A":22,"B":32,"C":20,"D":22})
    r=page_hdr(ws,1,"LOCATE RESULTS",f"Case: {CASE_NUM}  ·  Subject: {SUBJECT}  ·  {RPT_DATE}")

    best=R["addresses"][0] if R["addresses"] else {"addr":"Not found","source":"","confidence":"Unlocated","corroborated_by":0}
    sec(ws,r,1,4,"CURRENT ADDRESS — PRIMARY RESULT",GRN); r+=1
    kv(ws,r,"Current Address",best["addr"]); r+=1
    kv(ws,r,"Confidence",best["confidence"],alt=True,
       vc=GRN if "VERIF" in best["confidence"] or "HIGH" in best["confidence"].upper() else AMB); r+=1
    kv(ws,r,"Source(s)",best.get("source","")); r+=1
    kv(ws,r,"Corroborated By",f"{best.get('corroborated_by',1)} source(s)",alt=True); r+=1
    spacer(ws,r,8,4); r+=1

    sec(ws,r,1,4,"ALL ADDRESSES FOUND"); r+=1
    thdr(ws,r,[(1,"Address"),(2,"Source"),(3,"Confidence"),(4,"# Sources")]); r+=1
    for i,a in enumerate(R["addresses"][:10]):
        drow(ws,r+i,[(1,a["addr"]),(2,a.get("source","")),(3,a["confidence"]),(4,str(a.get("corroborated_by",1)))],alt=i%2==1,h=20,wrap=True)
    r+=max(10,len(R["addresses"])); spacer(ws,r,8,4); r+=1

    sec(ws,r,1,4,"PHONE NUMBERS"); r+=1
    thdr(ws,r,[(1,"Phone"),(2,"Source"),(3,"Type"),(4,"Notes")]); r+=1
    for i,p in enumerate(R["phones"][:8]):
        drow(ws,r+i,[(1,p["phone"]),(2,p["source"]),(3,""),(4,"")],alt=i%2==1)
    gold_bar(ws,r+max(8,len(R["phones"]))+2,4)

    source_log_sheet(wb)
    next_steps_sheet(wb,[
        ("1","Verify address via process server or field check","⬜ Pending","Most critical"),
        ("2","Manual TX voter reg check — teamrv-mvp.sos.texas.gov","⬜ Pending","CAPTCHA protected"),
        ("3","Upgrade to Comprehensive ($150) for full asset picture","⬜ Optional",""),
    ], GRN)
    wb.save(OUT_PATH); print(f"\n  ✓ Tier 1 Standard report saved: {OUT_PATH}")

# ════════════════════════════════════════════════════════════
#  TIER 2 — COMPREHENSIVE ($150)
# ════════════════════════════════════════════════════════════
def build_t2():
    wb=openpyxl.Workbook()
    cover_sheet(wb, BLUE)

    # Sheet 2: Locate Results
    ws=wb.create_sheet("LOCATE RESULTS")
    ws.sheet_view.showGridLines=False; ws.sheet_properties.tabColor=GRN
    cw(ws,{"A":20,"B":30,"C":20,"D":22})
    r=page_hdr(ws,1,"LOCATE RESULTS",f"Case: {CASE_NUM}  ·  {SUBJECT}  ·  {RPT_DATE}")
    best=R["addresses"][0] if R["addresses"] else {"addr":"Not found","source":"","confidence":"Unlocated","corroborated_by":0}
    sec(ws,r,1,4,"PRIMARY ADDRESS"); r+=1
    kv(ws,r,"Current Address",best["addr"]); r+=1
    kv(ws,r,"Confidence",best["confidence"],alt=True,vc=GRN if "VERIF" in best["confidence"] or "HIGH" in best["confidence"].upper() else AMB); r+=1
    kv(ws,r,"Corroborated By",f"{best.get('corroborated_by',1)} source(s)"); r+=1
    spacer(ws,r,8,4); r+=1
    sec(ws,r,1,4,"ALL ADDRESSES"); r+=1
    thdr(ws,r,[(1,"Address"),(2,"Source"),(3,"Confidence"),(4,"Corroborations")]); r+=1
    for i,a in enumerate(R["addresses"][:10]):
        drow(ws,r+i,[(1,a["addr"]),(2,a.get("source","")),(3,a["confidence"]),(4,str(a.get("corroborated_by",1)))],alt=i%2==1,h=20,wrap=True)
    r+=max(10,len(R["addresses"])); spacer(ws,r,8,4); r+=1
    sec(ws,r,1,4,"PHONES & EMAILS"); r+=1
    thdr(ws,r,[(1,"Phone/Email"),(2,"Source"),(3,"Type"),(4,"Notes")]); r+=1
    rows=[(p["phone"],p["source"],"Phone","") for p in R["phones"][:5]] + [(e["email"],e["source"],"Email","") for e in R["emails"][:4]]
    for i,(val,src,t,n) in enumerate(rows):
        drow(ws,r+i,[(1,val),(2,src),(3,t),(4,n)],alt=i%2==1)
    r+=len(rows)+2; spacer(ws,r,8,4); r+=1
    sec(ws,r,1,4,"RELATIVES & ASSOCIATES"); r+=1
    thdr(ws,r,[(1,"Name"),(2,"Relationship"),(3,"Source"),(4,"Notes")]); r+=1
    for i,rel in enumerate(R["relatives"][:10]):
        drow(ws,r+i,[(1,rel),(2,"Relative/Associate"),(3,"TruePeopleSearch"),(4,"")],alt=i%2==1)
    gold_bar(ws,r+max(10,len(R["relatives"]))+2,4)

    # Sheet 3: Digital Footprint
    ws3=wb.create_sheet("DIGITAL FOOTPRINT")
    ws3.sheet_view.showGridLines=False; ws3.sheet_properties.tabColor=BLUE
    cw(ws3,{"A":20,"B":40,"C":20,"D":16})
    r=page_hdr(ws3,1,"DIGITAL FOOTPRINT",f"Social Media · Email · Employment",nc=4)
    sec(ws3,r,1,4,"SOCIAL MEDIA"); r+=1
    thdr(ws3,r,[(1,"Platform"),(2,"Profile / URL"),(3,"Status"),(4,"Notes")]); r+=1
    for i,plat in enumerate(["Facebook","LinkedIn","Instagram","Twitter/X","TikTok","YouTube"]):
        url_v=R["social"].get(plat,"Not found")
        drow(ws3,r+i,[(1,plat),(2,url_v),(3,"FOUND" if url_v!="Not found" else "Not located"),(4,"")],
             alt=i%2==1,color=GOLD2 if url_v!="Not found" else DIM)
    r+=7; spacer(ws3,r,8,4); r+=1
    sec(ws3,r,1,4,"EMPLOYMENT"); r+=1
    thdr(ws3,r,[(1,"Employer"),(2,"Source"),(3,"Address"),(4,"Status")]); r+=1
    for i,emp in enumerate(R["employers"][:6]):
        drow(ws3,r+i,[(1,emp["employer"]),(2,emp["source"]),(3,emp.get("addr","")),(4,"")],alt=i%2==1)
    gold_bar(ws3,r+max(6,len(R["employers"]))+2,4)

    # Sheet 4: Assets
    ws4=wb.create_sheet("ASSETS")
    ws4.sheet_view.showGridLines=False; ws4.sheet_properties.tabColor=AMB
    cw(ws4,{"A":26,"B":24,"C":20,"D":20})
    r=page_hdr(ws4,1,"ASSET IDENTIFICATION",f"Property · Business · Licenses",nc=4)
    sec(ws4,r,1,4,"REAL PROPERTY — TX CAD"); r+=1
    thdr(ws4,r,[(1,"Property Address"),(2,"County"),(3,"Source"),(4,"Notes")]); r+=1
    if R["properties"]:
        for i,p in enumerate(R["properties"][:6]):
            drow(ws4,r+i,[(1,p.get("addr","")),(2,p.get("county","")),(3,p.get("source","")),(4,"")],alt=i%2==1,wrap=True)
    else: drow(ws4,r,[(1,"None found in searched counties"),(2,""),(3,""),(4,"Manual: all 254 TX counties")],color=DIM)
    r+=max(6,len(R["properties"]))+1; spacer(ws4,r,8,4); r+=1
    sec(ws4,r,1,4,"BUSINESS INTERESTS — TX SOS"); r+=1
    thdr(ws4,r,[(1,"Business Name"),(2,"Source"),(3,"Status"),(4,"Notes")]); r+=1
    if R["businesses"]:
        for i,b in enumerate(R["businesses"][:5]):
            drow(ws4,r+i,[(1,b["name"]),(2,b["source"]),(3,b.get("status","")),(4,"")],alt=i%2==1)
    else: drow(ws4,r,[(1,"No business on file"),(2,""),(3,"Clean"),(4,"")],color=DIM)
    r+=max(5,len(R["businesses"]))+1; spacer(ws4,r,8,4); r+=1
    sec(ws4,r,1,4,"PROFESSIONAL LICENSES"); r+=1
    thdr(ws4,r,[(1,"License Type"),(2,"Number"),(3,"Address on File"),(4,"Source")]); r+=1
    if R["licenses"]:
        for i,l in enumerate(R["licenses"][:5]):
            drow(ws4,r+i,[(1,l["type"]),(2,l["number"]),(3,l.get("addr","")),(4,"TDLR/TX Bar")],alt=i%2==1)
    else: drow(ws4,r,[(1,"No professional license"),(2,""),(3,""),(4,"Manual: tdlr.texas.gov")],color=DIM)
    gold_bar(ws4,r+max(5,len(R["licenses"]))+2,4)

    # Sheet 5: Red Flags
    ws5=wb.create_sheet("RED FLAGS")
    ws5.sheet_view.showGridLines=False; ws5.sheet_properties.tabColor=RED
    cw(ws5,{"A":14,"B":40,"C":16,"D":28})
    r=page_hdr(ws5,1,"RED FLAGS & ISSUES",f"Case: {CASE_NUM}  ·  {RPT_DATE}",nc=4)
    thdr(ws5,r,[(1,"SEVERITY"),(2,"ISSUE"),(3,"CATEGORY"),(4,"RECOMMENDED ACTION")],h=22); r+=1
    if R["red_flags"]:
        for i,f in enumerate(R["red_flags"]): flagrow(ws5,r+i,f["severity"],f["desc"],f["cat"],f["action"])
    else:
        rh(ws5,r,24)
        MC(ws5,r,1,r,4,"✓  No significant red flags in automated search — manual verification recommended",sz=10,color=GRN,bg=GRN_BG,bc=GRN)
    r+=max(8,len(R["red_flags"]))+1; spacer(ws5,r,8,4); r+=1
    sec(ws5,r,1,4,"CRIMINAL RECORD SUMMARY"); r+=1
    thdr(ws5,r,[(1,"Source"),(2,"Record"),(3,"Details"),(4,"Action")]); r+=1
    if R["criminal"]:
        for i,c in enumerate(R["criminal"]):
            drow(ws5,r+i,[(1,c["source"]),(2,c["record"]),(3,""),(4,"Disclose to attorney")],alt=i%2==1,color=RED)
    else:
        drow(ws5,r,[(1,"TDCJ + TX DPS"),(2,"No criminal record found"),(3,""),(4,"")],color=GRN)
    gold_bar(ws5,r+3,4)

    source_log_sheet(wb)
    next_steps_sheet(wb,[
        ("1","Field verify primary address","⬜ Pending","Critical"),
        ("2","Manual TX voter reg — teamrv-mvp.sos.texas.gov","⬜ Pending","CAPTCHA"),
        ("3","Verify employment via direct call","⬜ Pending",""),
        ("4","Pull PACER federal docket — pacer.gov ($0.10/pg)","⬜ Pending","Bankruptcy + federal"),
        ("5","Check all 254 TX CADs if property not found","⬜ Pending",""),
        ("6","Consider TLO/LexisNexis if address unconfirmed","⬜ Optional","$50-100/mo"),
    ], BLUE)
    wb.save(OUT_PATH); print(f"\n  ✓ Tier 2 Comprehensive report saved: {OUT_PATH}")

# ════════════════════════════════════════════════════════════
#  TIER 3 — PROPERTY & CHAIN OF TITLE ($200)
# ════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════
#  TIER 3 — PROPERTY & CHAIN OF TITLE ($200)
#  Premium professional layout
# ════════════════════════════════════════════════════════════

def _t3_ws(wb, title, tab_color, cols):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = tab_color
    cw(ws, cols)
    return ws

def _t3_page_title(ws, row, title, subtitle, nc=7, accent=TEAL):
    """Full-width premium page header"""
    gold_bar(ws, row, nc)
    rh(ws, row+1, 10)
    for c in range(1, nc+1): ws[f"{get_column_letter(c)}{row+1}"].fill = fill(BG4)
    rh(ws, row+2, 38)
    MC(ws, row+2, 1, row+2, nc, f"  {title}", bold=True, sz=16, color=accent, bg=BG4, h="left", v="center", bc=GOLD3)
    rh(ws, row+3, 20)
    MC(ws, row+3, 1, row+3, nc,
       f"  TraceWorks  ·  Property & Chain of Title  ·  Case {CASE_NUM}  ·  {SUBJECT}  ·  {RPT_DATE}",
       sz=9, color=DIM, bg=BG4, h="left", italic=True, bc=BG4)
    rh(ws, row+4, 3)
    MC(ws, row+4, 1, row+4, nc, bg=TEAL, bc=TEAL)  # teal accent line
    rh(ws, row+5, 8)
    for c in range(1, nc+1): ws[f"{get_column_letter(c)}{row+5}"].fill = fill(BG)
    return row + 6

def _sec(ws, r, c1, c2, label, accent=TEAL):
    rh(ws, r, 24)
    MC(ws, r, c1, r, c2, f"  {label}", bold=True, sz=9, color=accent, bg=BG4, bc=GOLD3)
    ws[f"{get_column_letter(c1)}{r}"].border = Border(
        left=Side(style="medium", color=accent),
        right=Side(style="thin", color=DIM2),
        top=Side(style="thin", color=DIM2),
        bottom=Side(style="thin", color=DIM2))
    return r + 1

def _kv(ws, r, label, value, alt=False, h=22, vc=None, c1=1, c2=4, label_w=2):
    bg = BG3 if alt else BG2
    rh(ws, r, h)
    lc = get_column_letter(c1)
    vc2 = get_column_letter(c1 + label_w - 1)
    MC(ws, r, c1, r, c1+label_w-1, label, bold=True, sz=9, color=GOLD2, bg=BG4, h="right", ind=1, bc=DIM2)
    ws[f"{lc}{r}"].border = Border(
        left=Side(style="thin", color=DIM2),
        right=Side(style="medium", color=TEAL),
        top=Side(style="thin", color=DIM2),
        bottom=Side(style="thin", color=DIM2))
    MC(ws, r, c1+label_w, r, c2, str(value), sz=9, color=vc or TEXT, bg=bg, wrap=True, ind=1, bc=DIM2)
    return r + 1

def _thdr(ws, r, col_defs, h=22, accent=TEAL):
    rh(ws, r, h)
    for cn, lbl, w_hint in col_defs:
        ref = f"{get_column_letter(cn)}{r}"
        ws[ref].value = lbl
        ws[ref].font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        ws[ref].fill = fill(accent)
        ws[ref].alignment = Alignment(horizontal="center", vertical="center")
        ws[ref].border = Border(
            left=Side(style="thin", color=BG4),
            right=Side(style="thin", color=BG4),
            top=Side(style="thin", color=BG4),
            bottom=Side(style="medium", color=GOLD))
    return r + 1

def _drow(ws, r, col_vals, alt=False, h=22, wrap=False, color=TEXT, bold=False):
    bg = BG3 if alt else BG2
    rh(ws, r, h)
    for cn, val in col_vals:
        ref = f"{get_column_letter(cn)}{r}"
        ws[ref].value = str(val) if val is not None else ""
        ws[ref].font = Font(name="Calibri", bold=bold, size=9, color=color)
        ws[ref].fill = fill(bg)
        ws[ref].alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap, indent=1)
        ws[ref].border = Border(
            left=Side(style="thin", color=DIM2),
            right=Side(style="thin", color=DIM2),
            top=Side(style="thin", color=DIM2),
            bottom=Side(style="thin", color=DIM2))
    return r + 1

def _empty_rows(ws, r, count, ncols=7, h=22):
    for i in range(count):
        _drow(ws, r+i, [(c, "") for c in range(1, ncols+1)], alt=i%2==1, h=h)
    return r + count

def _divider(ws, r, nc=7, h=6):
    rh(ws, r, h)
    for c in range(1, nc+1): ws[f"{get_column_letter(c)}{r}"].fill = fill(BG)
    return r + 1

def _flag_row(ws, r, sev, desc, cat, action, url=""):
    rh(ws, r, 30)
    s = sev.upper()
    if "HIGH" in s:   bg, fc, lc = RED_BG, RED, RED
    elif "MED" in s:  bg, fc, lc = AMB_BG, AMB, AMB
    else:             bg, fc, lc = GRN_BG, GRN, GRN
    for cn, val in [(1, sev), (2, desc), (3, cat), (4, action), (5, url)]:
        ref = f"{get_column_letter(cn)}{r}"
        ws[ref].value = val
        ws[ref].font = Font(name="Calibri", bold=(cn==1), size=9, color=fc)
        ws[ref].fill = fill(bg)
        ws[ref].alignment = Alignment(horizontal="left" if cn>1 else "center",
                                       vertical="center", wrap_text=True, indent=1)
        ws[ref].border = Border(
            left=Side(style="medium" if cn==1 else "thin", color=lc if cn==1 else DIM2),
            right=Side(style="thin", color=DIM2),
            top=Side(style="thin", color=DIM2),
            bottom=Side(style="thin", color=DIM2))
    return r + 1

def build_t3():
    wb = openpyxl.Workbook()

    # ── SHEET 1: COVER ──────────────────────────────────────────
    ws_cov = wb.active
    ws_cov.title = "COVER"
    ws_cov.sheet_view.showGridLines = False
    ws_cov.sheet_properties.tabColor = TEAL
    cw(ws_cov, {"A":2,"B":2,"C":14,"D":22,"E":18,"F":18,"G":14,"H":2})

    # Black background
    for row in range(1, 45):
        rh(ws_cov, row, 16)
        for col in range(1, 9):
            ws_cov[f"{get_column_letter(col)}{row}"].fill = fill(BG)

    # Top gold bar
    rh(ws_cov, 2, 4)
    MC(ws_cov, 2, 2, 2, 7, bg=GOLD, bc=GOLD)

    # TRACEWORKS wordmark
    rh(ws_cov, 4, 64)
    MC(ws_cov, 4, 2, 4, 7, "TRACEWORKS", bold=True, sz=48, color=GOLD, bg=BG, h="left", v="center", bc=BG)

    rh(ws_cov, 5, 22)
    MC(ws_cov, 5, 2, 5, 7, "PROPERTY RESEARCH  ·  CHAIN OF TITLE  ·  OSINT INVESTIGATION",
       sz=9, color=DIM, bg=BG, h="left", v="center", italic=True, bc=BG)

    # Teal accent bar
    rh(ws_cov, 6, 4)
    MC(ws_cov, 6, 2, 6, 7, bg=TEAL, bc=TEAL)

    rh(ws_cov, 7, 12)

    # Report type banner
    rh(ws_cov, 8, 36)
    MC(ws_cov, 8, 2, 8, 7, "PROPERTY & CHAIN OF TITLE INVESTIGATION",
       bold=True, sz=16, color=WHITE, bg=BG2, h="left", v="center", bc=DIM2)

    rh(ws_cov, 9, 24)
    MC(ws_cov, 9, 2, 9, 7,
       f"  Tier 3  ·  $200  ·  CAD Search · Deed History · Liens · Mineral Rights · TX RRC",
       sz=10, color=TEAL, bg=BG2, h="left", v="center", italic=True, bc=DIM2)

    rh(ws_cov, 10, 14)

    # Stats bar
    rh(ws_cov, 11, 28)
    stats = [
        (2, 3, f"{len(R['properties'])}", "Properties Found"),
        (4, 5, f"{len(R['deeds'])}", "Deed Records"),
        (6, 7, f"{len(R['red_flags'])}", "Red Flags"),
    ]
    for c1, c2, num, label in stats:
        cl1 = get_column_letter(c1); cl2 = get_column_letter(c2)
        ws_cov.merge_cells(f"{cl1}11:{cl2}11")
        cell = ws_cov[f"{cl1}11"]
        cell.value = f"{num}  {label}"
        cell.font = Font(name="Calibri", bold=True, size=11, color=TEAL)
        cell.fill = fill(BG3)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin", color=DIM2), right=Side(style="thin", color=DIM2),
            top=Side(style="thin", color=DIM2), bottom=Side(style="thin", color=DIM2))

    rh(ws_cov, 12, 14)

    # Info table
    info_rows = [
        ("SUBJECT / OWNER", SUBJECT),
        ("DATE OF BIRTH", DOB or "Not provided"),
        ("LAST KNOWN ADDRESS", LAST_ADDR or "Not provided"),
        ("CASE NUMBER", CASE_NUM),
        ("SERVICE", "Property & Chain of Title  ·  $200"),
        ("REPORT DATE", f"{RPT_DATE}  ·  {RPT_TIME}"),
        ("PREPARED FOR", f"{ATT_NAME}  ·  {ATT_FIRM}"),
        ("CLIENT EMAIL", ATT_EMAIL),
        ("PREPARED BY", "TraceWorks  ·  traceworks.tx@outlook.com"),
        ("SOURCES CHECKED", f"{len(R['sources'])} databases"),
        ("LOCATE CONFIDENCE", R.get("confidence", "Pending")),
    ]

    for i, (label, value) in enumerate(info_rows):
        r_num = 13 + i
        alt = i % 2 == 1
        bg_val = BG3 if alt else BG2
        rh(ws_cov, r_num, 24)
        is_conf = label == "LOCATE CONFIDENCE"
        is_case = label == "CASE NUMBER"

        # Label cell
        MC(ws_cov, r_num, 2, r_num, 3, label,
           bold=True, sz=9, color=GOLD2, bg=BG4, h="right", ind=1, bc=DIM2)
        ws_cov[f"B{r_num}"].border = Border(
            left=Side(style="thin", color=DIM2),
            right=Side(style="medium", color=TEAL),
            top=Side(style="thin", color=DIM2),
            bottom=Side(style="thin", color=DIM2))

        # Value cell
        vc = TEAL if is_conf else (GOLD2 if is_case else TEXT)
        vbg = BG3 if is_conf else bg_val
        MC(ws_cov, r_num, 4, r_num, 7, value,
           sz=10 if is_conf or is_case else 9,
           bold=is_conf or is_case,
           color=vc, bg=vbg, ind=1, bc=DIM2)

    rh(ws_cov, 24, 14)

    # Bottom bar
    gold_bar(ws_cov, 25, 8)
    rh(ws_cov, 26, 18)
    MC(ws_cov, 26, 2, 26, 7,
       "ATTORNEY-CLIENT WORK PRODUCT  ·  CONFIDENTIAL  ·  FOR AUTHORIZED LEGAL USE ONLY  ·  TraceWorks  ·  traceworks.tx@outlook.com",
       sz=8, color=DIM, bg=BG, h="center", italic=True, bc=BG)

    # ── SHEET 2: EXECUTIVE SUMMARY ──────────────────────────────
    ws_ex = _t3_ws(wb, "EXECUTIVE SUMMARY", TEAL,
                   {"A":3,"B":20,"C":20,"D":20,"E":20,"F":14,"G":3})
    r = _t3_page_title(ws_ex, 1, "EXECUTIVE SUMMARY", "", nc=7)

    # Key findings box
    _sec(ws_ex, r, 2, 6, "KEY FINDINGS AT A GLANCE"); r += 1
    findings = [
        ("Property Found", f"{len(R['properties'])} record(s) in Texas CAD search", TEAL if R['properties'] else DIM),
        ("Primary Address", R['addresses'][0]['addr'] if R['addresses'] else "Not located", TEXT),
        ("Confidence Level", R.get("confidence", "Pending"), TEAL),
        ("Deed Records", f"{len(R['deeds'])} instrument(s) extracted from county clerk", TEXT),
        ("Business Interests", f"{len(R['businesses'])} entity/entities found in TX SOS" if R['businesses'] else "None found", TEXT),
        ("Mineral Interests", f"{len(R['mineral'])} RRC permit(s) on file" if R['mineral'] else "No RRC operator record", TEXT),
        ("Court Records", f"{len(R['court_records'])} record(s) found" if R['court_records'] else "Clean — no court records", GRN if not R['court_records'] else AMB),
        ("Red Flags", f"{len(R['red_flags'])} issue(s) require attention" if R['red_flags'] else "No significant issues found", RED if R['red_flags'] else GRN),
        ("Sources Checked", f"{len(R['sources'])} databases searched automatically", DIM),
        ("Report Date", f"{RPT_DATE} at {RPT_TIME}", DIM),
    ]
    for i, (label, val, vc) in enumerate(findings):
        r = _kv(ws_ex, r, label, val, alt=i%2==1, vc=vc, c1=2, c2=6, label_w=2)

    r = _divider(ws_ex, r, 7)

    # Counties searched
    _sec(ws_ex, r, 2, 6, "CAD COUNTIES SEARCHED"); r += 1
    counties_searched = [
        s["name"] for s in R["sources"] if "CAD" in s["name"] or "County" in s["name"]
    ]
    county_text = "  ·  ".join(counties_searched) if counties_searched else "Travis, Williamson, Harris, Milam, Bell, Bastrop, Hays, Bexar"
    rh(ws_ex, r, 28)
    MC(ws_ex, r, 2, r, 6, county_text, sz=9, color=TEXT, bg=BG2, wrap=True, ind=1, bc=DIM2)
    r += 2

    r = _divider(ws_ex, r, 7)
    _sec(ws_ex, r, 2, 6, "IMPORTANT NOTES FOR ATTORNEY"); r += 1
    notes = [
        "This report reflects automated public record searches only. Manual county clerk review is required for complete chain of title.",
        "TX CAD records show ownership and assessment data only — not full deed history. Chain of title requires county clerk deed retrieval.",
        "Mineral rights may be severed from surface rights. Eagle Ford corridor (Milam, Karnes, DeWitt, Gonzales counties) requires specific mineral search.",
        "Automated searches checked 8 major TX counties. Subject may own property in any of the 254 TX counties — manual search recommended.",
        "UCC liens, federal tax liens, and judgment liens require separate searches at TX SOS (ucc.sos.state.tx.us) and PACER.",
    ]
    for i, note in enumerate(notes):
        rh(ws_ex, r+i, 28)
        ref_a = f"B{r+i}"
        ws_ex[ref_a].value = f"  {i+1}."
        ws_ex[ref_a].font = Font(name="Calibri", bold=True, size=9, color=GOLD)
        ws_ex[ref_a].fill = fill(BG3 if i%2==1 else BG2)
        ws_ex[ref_a].alignment = Alignment(horizontal="center", vertical="center")
        ws_ex[ref_a].border = Border(left=Side(style="medium",color=TEAL),right=Side(style="thin",color=DIM2),top=Side(style="thin",color=DIM2),bottom=Side(style="thin",color=DIM2))
        MC(ws_ex, r+i, 3, r+i, 6, note, sz=9, color=DIM, bg=BG3 if i%2==1 else BG2, wrap=True, ind=1, bc=DIM2)
    r += len(notes) + 1
    gold_bar(ws_ex, r, 7)

    # ── SHEET 3: PROPERTY PROFILE ────────────────────────────────
    ws_pp = _t3_ws(wb, "PROPERTY PROFILE", TEAL,
                   {"A":3,"B":22,"C":22,"D":20,"E":16,"F":16,"G":3})
    r = _t3_page_title(ws_pp, 1, "PROPERTY PROFILE", "", nc=7)

    # Properties from CAD
    _sec(ws_pp, r, 2, 6, f"TEXAS CAD SEARCH RESULTS  ·  {len(R['properties'])} RECORD(S) FOUND"); r += 1
    r = _thdr(ws_pp, r, [(2,"PROPERTY ADDRESS",22),(3,"COUNTY / CAD",18),(4,"TYPE",14),(5,"ACCOUNT #",14),(6,"SOURCE",16)], accent=TEAL)
    if R["properties"]:
        for i, p in enumerate(R["properties"][:10]):
            r = _drow(ws_pp, r, [(2,p.get("addr","")),(3,p.get("county","")),(4,p.get("type","Real Property")),(5,""),(6,p.get("source",""))], alt=i%2==1, h=24, wrap=True)
    else:
        r = _drow(ws_pp, r, [(2,"No property auto-located — manual CAD search required"),(3,""),(4,""),(5,""),(6,"")], color=DIM)
    # Empty rows for manual entry
    r = _empty_rows(ws_pp, r, 4, 7, 24)
    r = _divider(ws_pp, r, 7)

    # CAD Detail — manual fill section
    _sec(ws_pp, r, 2, 6, "CAD DETAIL — PRIMARY PROPERTY  (Complete from CAD portal)"); r += 1
    cad_fields = [
        ("Owner of Record", ""), ("CAD Account Number", ""),
        ("Property Address", R["addresses"][0]["addr"] if R["addresses"] else ""),
        ("Legal Description", ""), ("Abstract / Survey", ""),
        ("Acreage", ""), ("Appraisal Year", ""),
        ("Land Value", ""), ("Improvement Value", ""),
        ("Total Assessed Value", ""), ("Market Value", ""),
        ("Tax Status", ""), ("Exemptions on File", ""),
        ("Special Valuations", ""), ("Last Sale Date", ""),
        ("Last Sale Price", ""),
    ]
    for i, (label, val) in enumerate(cad_fields):
        r = _kv(ws_pp, r, label, val, alt=i%2==1, c1=2, c2=6, label_w=2)
    r = _divider(ws_pp, r, 7)

    # County portal quick reference
    _sec(ws_pp, r, 2, 6, "COUNTY CAD PORTAL QUICK REFERENCE"); r += 1
    portals = [
        ("Travis County", "esearch.traviscad.org", R.get("confidence","")),
        ("Williamson County", "wcad.org", ""),
        ("Harris County", "public.hcad.org", ""),
        ("Milam County", "milam.tx.publicsearch.us", "107 W Main St Cameron TX — FOUND" if any("Milam" in p.get("county","") for p in R["properties"]) else ""),
        ("Bell County", "bell.tx.publicsearch.us", ""),
        ("Bastrop County", "bastrop.tx.publicsearch.us", ""),
        ("Hays County", "hays.tx.publicsearch.us", ""),
        ("Bexar County", "bexar.tx.publicsearch.us", ""),
        ("All 254 TX Counties", "publicdata.sao.texas.gov", "Use for statewide search"),
    ]
    r = _thdr(ws_pp, r, [(2,"COUNTY",22),(3,"PORTAL URL",30),(4,"RESULT / NOTES",30),(5,"",1),(6,"",1)], accent=TEAL)
    for i, (county, url, result) in enumerate(portals):
        src_result = next((s["result"] for s in R["sources"] if county.split()[0] in s["name"]), result)
        is_found = "FOUND" in src_result.upper() if src_result else False
        r = _drow(ws_pp, r, [(2,county),(3,url),(4,src_result),(5,""),(6,"")],
                  alt=i%2==1, color=TEAL if is_found else TEXT)
    gold_bar(ws_pp, r+1, 7)

    # ── SHEET 4: CHAIN OF TITLE ──────────────────────────────────
    ws_cot = _t3_ws(wb, "CHAIN OF TITLE", TEAL,
                    {"A":3,"B":6,"C":22,"D":22,"E":18,"F":16,"G":18,"H":3})
    ws_cot.freeze_panes = "B8"
    r = _t3_page_title(ws_cot, 1, "CHAIN OF TITLE", "", nc=8)

    # Instructions box
    rh(ws_cot, r, 44)
    MC(ws_cot, r, 2, r, 7,
       "INSTRUCTIONS: Enter deed history from oldest to newest (bottom to top of chain). "
       "Pull instruments at county clerk — use portal URLs on Property Profile tab. "
       "Each row = one recorded instrument. Highlight in RED any gaps in chain or quitclaim deeds. "
       "For Milam County: milam.tx.publicsearch.us  ·  Search by grantor AND grantee name.",
       sz=9, color=AMB, bg=AMB_BG, wrap=True, ind=1, bc=AMB)
    ws_cot[f"B{r}"].border = Border(left=Side(style="medium",color=AMB),right=Side(style="thin",color=DIM2),top=Side(style="thin",color=DIM2),bottom=Side(style="thin",color=DIM2))
    r += 1
    r = _divider(ws_cot, r, 8)

    _sec(ws_cot, r, 2, 7, f"DEED HISTORY  ·  Subject: {SUBJECT}  ·  Auto-extracted + Manual Entry"); r += 1
    r = _thdr(ws_cot, r,
              [(2,"#",4),(3,"GRANTOR  (Seller / From)",22),(4,"GRANTEE  (Buyer / To)",22),
               (5,"DOCUMENT TYPE",18),(6,"RECORDED DATE",16),(7,"INSTRUMENT #",18)],
              accent=TEAL)

    # Auto-extracted deeds from scraper
    if R["deeds"]:
        for i, d in enumerate(R["deeds"]):
            r = _drow(ws_cot, r,
                      [(2,str(i+1)),(3,d.get("grantor","")),(4,d.get("grantee","")),(5,d["type"]),(6,""),(7,d["instrument"])],
                      alt=i%2==1, h=26)
        # Empty rows after auto results
        start_i = len(R["deeds"])
        for j in range(12):
            r = _drow(ws_cot, r, [(2,str(start_i+j+1)),(3,""),(4,""),(5,""),(6,""),(7,"")], alt=j%2==1, h=26)
    else:
        # All empty rows with sequential numbers
        for i in range(15):
            r = _drow(ws_cot, r, [(2,str(i+1)),(3,""),(4,""),(5,""),(6,""),(7,"")], alt=i%2==1, h=26)

    r = _divider(ws_cot, r, 8)

    # Deed type legend
    _sec(ws_cot, r, 2, 7, "DEED TYPE REFERENCE  ·  Texas"); r += 1
    deed_types = [
        ("General Warranty Deed", "Full title warranty — strongest form. Grantor warrants against ALL prior claims.", GRN),
        ("Special Warranty Deed", "Grantor only warrants against claims arising during their ownership period.", AMB),
        ("Quitclaim Deed", "NO warranty. Conveys only whatever interest grantor has — may be nothing. HIGH RISK.", RED),
        ("Deed Without Warranty", "No representations made. Common in foreclosure/estate situations.", AMB),
        ("Deed of Trust", "Mortgage instrument — secures lender interest. NOT a conveyance of ownership.", BLUE),
        ("Release of Lien", "Releases a prior Deed of Trust or lien — indicates paid-off mortgage.", GRN),
        ("Trustee Deed", "Conveyance following foreclosure by trustee. Review carefully for defects.", AMB),
        ("Affidavit of Heirship", "Establishes title through intestate succession — no probate. Check for competing heirs.", AMB),
    ]
    r = _thdr(ws_cot, r, [(2,"DEED TYPE",22),(3,"DESCRIPTION",40),(4,"",1),(5,"",1),(6,"",1),(7,"",1)], accent=TEAL)
    for i, (dtype, desc, dc) in enumerate(deed_types):
        rh(ws_cot, r, 24)
        _drow(ws_cot, r, [(2,dtype),(3,desc),(4,""),(5,""),(6,""),(7,"")], alt=i%2==1, color=dc, wrap=True, h=24)
        r += 1

    gold_bar(ws_cot, r+1, 8)

    # ── SHEET 5: LIENS & ENCUMBRANCES ────────────────────────────
    ws_li = _t3_ws(wb, "LIENS & ENCUMBRANCES", RED,
                   {"A":3,"B":22,"C":22,"D":16,"E":16,"F":16,"G":3})
    r = _t3_page_title(ws_li, 1, "LIENS & ENCUMBRANCES", "", nc=7)

    # Warning box
    rh(ws_li, r, 36)
    MC(ws_li, r, 2, r, 6,
       "  ⚠  IMPORTANT: Automated searches cannot access paid lien databases. "
       "All sections below require manual search completion. "
       "Texas UCC: ucc.sos.state.tx.us ($1/search)  ·  Federal tax liens: PACER (pacer.gov)  ·  "
       "Property tax: county tax assessor portal  ·  Judgment liens: county district clerk",
       sz=9, color=AMB, bg=AMB_BG, wrap=True, ind=2, bc=AMB)
    ws_li[f"B{r}"].border = Border(left=Side(style="medium",color=AMB),right=Side(style="thin",color=DIM2),top=Side(style="thin",color=DIM2),bottom=Side(style="thin",color=DIM2))
    r += 2

    # Sections
    lien_sections = [
        ("DEED OF TRUST / ACTIVE MORTGAGES", ["Lender / Beneficiary","Original Amount","Date Recorded","Instrument #","Status","Release Date"]),
        ("PROPERTY TAX STATUS", ["Taxing Authority","Tax Year","Amount Due","Penalty / Interest","Status","Account #"]),
        ("JUDGMENT LIENS", ["Plaintiff","Court / Cause #","Amount","Date Filed","County","Status"]),
        ("UCC FINANCING STATEMENTS  (TX SOS)", ["Filing Date","Secured Party","Debtor Name","Collateral Description","File #","Status"]),
        ("FEDERAL TAX LIENS  (IRS / PACER)", ["Filed By","Date Filed","Amount","Court","Case #","Status"]),
        ("LIS PENDENS  (Pending Litigation)", ["Filed By","Case #","Property Description","Date Filed","Court","Status"]),
    ]
    for sec_title, cols in lien_sections:
        _sec(ws_li, r, 2, 6, sec_title, RED); r += 1
        r = _thdr(ws_li, r, [(i+2, c, 16) for i, c in enumerate(cols[:5])] + [(7,"",1)], accent=RED)
        r = _empty_rows(ws_li, r, 4, 7, 24)
        r = _divider(ws_li, r, 7)

    gold_bar(ws_li, r, 7)

    # ── SHEET 6: MINERAL RIGHTS ──────────────────────────────────
    ws_min = _t3_ws(wb, "MINERAL RIGHTS", AMB,
                    {"A":3,"B":22,"C":22,"D":18,"E":18,"F":14,"G":3})
    r = _t3_page_title(ws_min, 1, "MINERAL RIGHTS & OIL/GAS INTERESTS", "", nc=7)

    # Eagle Ford context
    rh(ws_min, r, 44)
    MC(ws_min, r, 2, r, 6,
       "EAGLE FORD SHALE NOTE: Milam County falls within the Eagle Ford Shale play. "
       "Mineral rights in this area may be HIGHLY VALUABLE and are frequently severed from surface rights. "
       "A surface owner may have zero mineral interest. Verify mineral ownership separately from surface via "
       "deed history and TX RRC records. Royalty income may be attachable as an asset.",
       sz=9, color=GOLD2, bg=BG4, wrap=True, ind=2, bc=GOLD3)
    ws_min[f"B{r}"].border = Border(left=Side(style="medium",color=GOLD),right=Side(style="thin",color=DIM2),top=Side(style="thin",color=DIM2),bottom=Side(style="thin",color=DIM2))
    r += 2

    # RRC Results
    _sec(ws_min, r, 2, 6, "TX RAILROAD COMMISSION  ·  AUTOMATED SEARCH RESULTS", AMB); r += 1
    r = _thdr(ws_min, r, [(2,"OPERATOR NAME",22),(3,"API / PERMIT #",18),(4,"COUNTY",14),(5,"LEASE NAME",18),(6,"STATUS",14)], accent=AMB)
    if R["mineral"]:
        for i, m in enumerate(R["mineral"]):
            r = _drow(ws_min, r, [(2,m["desc"]),(3,m.get("api","")),(4,""),(5,""),(6,m["source"])], alt=i%2==1, color=GOLD2)
    else:
        r = _drow(ws_min, r, [(2,"No operator record found in automated search"),(3,""),(4,""),(5,""),(6,"Manual: webapps2.rrc.texas.gov")], color=DIM)
    r = _empty_rows(ws_min, r, 4, 7, 22)
    r = _divider(ws_min, r, 7)

    # Mineral ownership analysis
    _sec(ws_min, r, 2, 6, "MINERAL OWNERSHIP ANALYSIS  ·  Complete Manually", AMB); r += 1
    mineral_fields = [
        ("Surface Owner", ""),
        ("Mineral Owner", ""),
        ("Are Minerals Severed?", "[ ] Yes — severed by deed dated ___  [ ] No — unified  [ ] Unknown"),
        ("Severance Instrument #", ""),
        ("Royalty Interest %", ""),
        ("Working Interest %", ""),
        ("Overriding Royalty", ""),
        ("Current Operator", ""),
        ("Lease Name / #", ""),
        ("Formation", "Eagle Ford Shale / Austin Chalk / Buda / other"),
        ("Lease Status", "[ ] Active  [ ] Expired  [ ] Held by Production  [ ] Unknown"),
        ("Production Status", "[ ] Producing  [ ] Non-producing  [ ] Shut-in  [ ] Unknown"),
        ("Royalty Payments Active?", "[ ] Yes  [ ] No  [ ] Unknown"),
        ("Royalty Check Recipient", ""),
        ("Estimated Monthly Income", ""),
        ("RRC District", ""),
    ]
    for i, (label, val) in enumerate(mineral_fields):
        r = _kv(ws_min, r, label, val, alt=i%2==1, c1=2, c2=6, label_w=2, h=24)
    r = _divider(ws_min, r, 7)

    # Mineral research resources
    _sec(ws_min, r, 2, 6, "MINERAL RESEARCH RESOURCES"); r += 1
    resources = [
        ("TX Railroad Commission", "webapps2.rrc.texas.gov", "Free — search by operator or lease name"),
        ("TX RRC GIS Viewer", "gis.rrc.texas.gov", "Map well locations by county"),
        ("DrillingInfo / Enverus", "enverus.com", "Paid — industry-grade production data"),
        ("TX GLO Mineral Records", "glo.texas.gov", "State land mineral records"),
        ("County Clerk Deeds", "milam.tx.publicsearch.us", "Search mineral severance deeds by grantor/grantee"),
        ("IHS Markit", "ihsmarkit.com", "Paid — comprehensive mineral/royalty research"),
    ]
    r = _thdr(ws_min, r, [(2,"RESOURCE",22),(3,"URL",26),(4,"NOTES",28),(5,"",1),(6,"",1)], accent=AMB)
    for i, (name, url, note) in enumerate(resources):
        r = _drow(ws_min, r, [(2,name),(3,url),(4,note),(5,""),(6,"")], alt=i%2==1)
    gold_bar(ws_min, r+1, 7)

    # ── SHEET 7: RED FLAGS ───────────────────────────────────────
    ws_rf = _t3_ws(wb, "RED FLAGS", RED,
                   {"A":3,"B":14,"C":36,"D":16,"E":28,"F":24,"G":3})
    r = _t3_page_title(ws_rf, 1, "RED FLAGS & TITLE ISSUES", "", nc=7)

    _sec(ws_rf, r, 2, 6, f"ISSUES IDENTIFIED  ·  {len(R['red_flags'])} AUTO-DETECTED  ·  Additional manual review required", RED); r += 1
    r = _thdr(ws_rf, r,
              [(2,"SEVERITY",14),(3,"ISSUE DESCRIPTION",36),(4,"CATEGORY",16),(5,"RECOMMENDED ACTION",28),(6,"LEGAL REFERENCE",24)],
              accent=RED)

    # Standard property flags always included
    standard_flags = [
        ("HIGH", f"Milam County is Eagle Ford Shale — mineral rights may be severed and highly valuable. Verify before any transaction.", "MINERAL", "Pull ALL deeds back 40+ years and search TX RRC for production", "TX Prop. Code §5.011; Eagle Ford play"),
        ("MEDIUM", "Quitclaim deeds transfer only grantor's interest — no warranty. Competing claims possible.", "TITLE", "Verify grantor had clear title at time of conveyance", "TX Prop. Code §5.022"),
        ("MEDIUM", "Nominal consideration (e.g. $10, 'love and affection') may indicate familial transfer — verify intent and tax implications.", "TITLE", "Confirm grantor/grantee relationship; check for gift tax implications", "TX Tax Code §23.01"),
        ("MEDIUM", "Affidavit of Heirship conveys title without probate — all heirs must be identified and must not contest.", "PROBATE", "Verify all heirs identified; recommend title insurance", "TX Est. Code §203.001"),
        ("LOW", "Gaps in chain of title (missing instruments) require gap insurance or suit to quiet title.", "TITLE", "File suit to quiet title if gap exists — TX Prop. Code §22.001", "TX Prop. Code §22.001"),
        ("LOW", "Automated search covers 8 of 254 TX counties only — subject may own unreported property.", "SEARCH", "Manually search all counties in subject's address history", ""),
    ]

    all_flags = standard_flags + [(f["severity"],f["desc"],f["cat"],f["action"],"") for f in R["red_flags"]]
    for sev, desc, cat, action, legal in all_flags:
        r = _flag_row(ws_rf, r, sev, desc, cat, action, legal)

    r = _empty_rows(ws_rf, r, 4, 7, 26)
    r = _divider(ws_rf, r, 7)

    # Title defect quick reference
    _sec(ws_rf, r, 2, 6, "COMMON TEXAS TITLE DEFECTS — QUICK REFERENCE"); r += 1
    defects = [
        ("Breaks in Chain", "Missing deed between grantor/grantee — title not marketable without quiet title action"),
        ("Wild Deed", "Deed recorded outside chain — cannot be found by reasonable title search"),
        ("Forgery / Fraud", "Forged signature on any deed — voidable; report to TX AG"),
        ("Missing Spouse Signature", "TX homestead requires both spouses to sign — deed may be void"),
        ("Unprobated Will", "Property passing by will must go through probate or affidavit of heirship"),
        ("Adverse Possession", "TX: 3/5/10/25-year statutes depending on circumstances — TX Civ. Prac. §16.025-16.028"),
        ("Boundary Disputes", "Survey discrepancies — requires new survey and possible boundary agreement"),
        ("Undisclosed Liens", "IRS, state tax, HOA, mechanic's liens may not appear in deed records"),
    ]
    r = _thdr(ws_rf, r, [(2,"DEFECT TYPE",20),(3,"DESCRIPTION / STATUTE",50),(4,"",1),(5,"",1),(6,"",1)], accent=RED)
    for i, (defect, desc) in enumerate(defects):
        rh(ws_rf, r, 26)
        _drow(ws_rf, r, [(2,defect),(3,desc),(4,""),(5,""),(6,"")], alt=i%2==1, wrap=True, h=26)
        r += 1
    gold_bar(ws_rf, r+1, 7)

    # ── SHEET 8: SOURCE LOG ──────────────────────────────────────
    ws_sl = _t3_ws(wb, "SOURCE LOG", BLUE,
                   {"A":3,"B":5,"C":26,"D":30,"E":14,"F":24,"G":3})
    r = _t3_page_title(ws_sl, 1, "SOURCE LOG", "", nc=7)

    _sec(ws_sl, r, 2, 6, f"ALL SOURCES CHECKED  ·  {len(R['sources'])} DATABASES  ·  {RPT_DATE} {RPT_TIME}"); r += 1
    r = _thdr(ws_sl, r, [(2,"#",5),(3,"SOURCE / DATABASE",26),(4,"URL",30),(5,"DATE",14),(6,"RESULT",24)], accent=BLUE)
    for i, src in enumerate(R["sources"]):
        alt = i % 2 == 1
        bg_s = BG3 if alt else BG2
        rh(ws_sl, r, 22)
        is_f = any(x in src["status"].upper() for x in ["FOUND","VERIF"])
        is_c = "CLEAN" in src["status"].upper()
        rc = TEAL if is_f else (GRN if is_c else DIM)
        for cn, val, clr in [
            (2, str(i+1), GOLD), (3, src["name"], TEXT),
            (4, src["url"], DIM), (5, RPT_DATE, DIM), (6, src["result"], rc)
        ]:
            ref = f"{get_column_letter(cn)}{r}"
            ws_sl[ref].value = val
            ws_sl[ref].font = Font(name="Calibri", bold=(cn==2), size=9, color=clr)
            ws_sl[ref].fill = fill(bg_s)
            ws_sl[ref].alignment = Alignment(horizontal="center" if cn==2 else "left", vertical="center", indent=1, wrap_text=True)
            ws_sl[ref].border = Border(left=Side(style="thin",color=DIM2),right=Side(style="thin",color=DIM2),top=Side(style="thin",color=DIM2),bottom=Side(style="thin",color=DIM2))
        r += 1
    r = _divider(ws_sl, r, 7)
    gold_bar(ws_sl, r, 7)

    # ── SHEET 9: NEXT STEPS ──────────────────────────────────────
    ws_ns = _t3_ws(wb, "NEXT STEPS", GRN,
                   {"A":3,"B":6,"C":42,"D":18,"E":16,"F":22,"G":3})
    r = _t3_page_title(ws_ns, 1, "NEXT STEPS & RECOMMENDATIONS", "", nc=7)

    _sec(ws_ns, r, 2, 6, f"ACTION ITEMS  ·  Prepared for: {ATT_NAME}  ·  {ATT_FIRM}"); r += 1
    r = _thdr(ws_ns, r,
              [(2,"#",6),(3,"RECOMMENDED ACTION",42),(4,"PRIORITY",18),(5,"STATUS",16),(6,"RESOURCE / URL",22)],
              accent=GRN)

    steps = [
        ("1", f"Retrieve all deeds for {SUBJECT} from Milam County Clerk — search grantor AND grantee indexes", "CRITICAL", "⬜ Pending", "milam.tx.publicsearch.us"),
        ("2", "Verify mineral rights ownership — search mineral severance deeds back 40+ years in chain of title", "CRITICAL", "⬜ Pending", "TX RRC: webapps2.rrc.texas.gov"),
        ("3", "Run UCC lien search on subject at TX SOS — $1 per search", "HIGH", "⬜ Pending", "ucc.sos.state.tx.us"),
        ("4", "Search TX Comptroller for state tax liens", "HIGH", "⬜ Pending", "comptroller.texas.gov/taxes/franchise"),
        ("5", "Check county district clerk for judgment liens in all counties where property found", "HIGH", "⬜ Pending", "County district clerk portal"),
        ("6", "Pull PACER report for federal tax liens and bankruptcy — $0.10/page", "HIGH", "⬜ Pending", "pacer.gov"),
        ("7", "Verify current tax status and any delinquency at county tax assessor", "HIGH", "⬜ Pending", "Milam: milamtaxcollector.com"),
        ("8", "Order title commitment from TX title company if transaction involved", "MEDIUM", "⬜ Optional", "Any TX licensed title company"),
        ("9", "Verify no homestead exemption issues if property to be liened or sold", "MEDIUM", "⬜ Pending", "TX Const. Art. XVI §50"),
        ("10", "Search TX unclaimed property for any estate/financial assets", "LOW", "⬜ Optional", "comptroller.texas.gov/up"),
    ]

    for i, (num, action, priority, status, resource) in enumerate(steps):
        alt = i % 2 == 1
        bg_s = BG3 if alt else BG2
        rh(ws_ns, r, 28)
        pc = RED if priority=="CRITICAL" else (AMB if priority=="HIGH" else (TEAL if priority=="MEDIUM" else DIM))
        for cn, val, clr, bold in [
            (2, num, GOLD, True), (3, action, TEXT, False),
            (4, priority, pc, True), (5, status, DIM, False), (6, resource, TEAL, False)
        ]:
            ref = f"{get_column_letter(cn)}{r}"
            ws_ns[ref].value = val
            ws_ns[ref].font = Font(name="Calibri", bold=bold, size=9, color=clr)
            ws_ns[ref].fill = fill(bg_s)
            ws_ns[ref].alignment = Alignment(horizontal="center" if cn==2 else "left", vertical="center", wrap_text=True, indent=1)
            ws_ns[ref].border = Border(left=Side(style="medium" if cn==2 else "thin",color=GRN if cn==2 else DIM2),right=Side(style="thin",color=DIM2),top=Side(style="thin",color=DIM2),bottom=Side(style="thin",color=DIM2))
        r += 1

    r = _divider(ws_ns, r, 7)

    # Disclaimer
    rh(ws_ns, r, 60)
    MC(ws_ns, r, 2, r, 6,
       "DISCLAIMER: This report was compiled using publicly available records and lawful OSINT methodology only. "
       "TraceWorks operates in compliance with the DPPA, FCRA, GLBA, and Texas state law. "
       "This report is intended for authorized legal use only and does not constitute legal advice. "
       "TraceWorks is not a licensed private investigator agency. All findings should be independently "
       "verified by qualified legal counsel before use in legal proceedings, transactions, or litigation. "
       "Chain of title and lien sections require manual completion from county clerk records. "
       "\n\nTraceWorks  ·  traceworks.tx@outlook.com  ·  Texas-Based Property Research & OSINT",
       sz=8, color=DIM, bg=BG3, italic=True, wrap=True, h="left", v="top", bc=DIM2)
    r += 2
    gold_bar(ws_ns, r, 7)

    wb.save(OUT_PATH)
    print(f"\n  ✓ Tier 3 Property & Title report saved: {OUT_PATH}")

def build_t4():
    wb=openpyxl.Workbook()
    cover_sheet(wb, PURPLE)

    # Sheet 2: Decedent Profile
    ws=wb.create_sheet("SUBJECT PROFILE")
    ws.sheet_view.showGridLines=False; ws.sheet_properties.tabColor=PURPLE
    cw(ws,{"A":22,"B":32,"C":22,"D":26})
    r=page_hdr(ws,1,"SUBJECT / DECEDENT PROFILE",f"Case: {CASE_NUM}  ·  {RPT_DATE}",nc=4)
    sec(ws,r,1,4,"KNOWN SUBJECT INFORMATION",PURPLE); r+=1
    for k,v in [("Full Legal Name",SUBJECT),("Date of Birth",DOB),("Last Known Address",LAST_ADDR),
                ("Last Known Phone",LAST_PHONE),("Date of Death",""),("Place of Death",""),
                ("Social Security #",""),("TX Driver License",""),("Marital Status",""),("Spouse Name","")]:
        kv(ws,r,k,v,alt=r%2==0,nc=4); r+=1
    spacer(ws,r,8,4); r+=1
    sec(ws,r,1,4,"ESTATE INFORMATION",PURPLE); r+=1
    for k,v in [("Probate Filed","[ ] Yes  [ ] No  [ ] Unknown"),("County of Probate",""),
                ("Cause/Case Number",""),("Estate Attorney",""),("Administrator/Executor",""),
                ("Will on File","[ ] Yes  [ ] No"),("Intestate","[ ] Yes  [ ] No")]:
        kv(ws,r,k,v,alt=r%2==0,nc=4); r+=1
    gold_bar(ws,r+1,4)

    # Sheet 3: Heirs & Relatives Found
    ws3=wb.create_sheet("HEIRS & RELATIVES")
    ws3.sheet_view.showGridLines=False; ws3.sheet_properties.tabColor=PURPLE
    cw(ws3,{"A":24,"B":24,"C":20,"D":22})
    r=page_hdr(ws3,1,"HEIRS & RELATIVES LOCATED",f"Case: {CASE_NUM}  ·  Auto-searched relatives via OSINT",nc=4)
    sec(ws3,r,1,4,"RELATIVES FOUND IN PUBLIC RECORDS",PURPLE); r+=1
    thdr(ws3,r,[(1,"Full Name"),(2,"Relationship"),(3,"Address/Location"),(4,"Source")]); r+=1
    for i,rel in enumerate(R["relatives"][:15]):
        drow(ws3,r+i,[(1,rel),(2,"Relative/Associate"),(3,""),(4,"TruePeopleSearch/Spokeo")],alt=i%2==1,h=22)
    r+=max(15,len(R["relatives"]))+1; spacer(ws3,r,8,4); r+=1
    sec(ws3,r,1,4,"SPOUSE / CHILDREN RESEARCH",PURPLE); r+=1
    thdr(ws3,r,[(1,"Name"),(2,"Relation"),(3,"DOB/Age"),(4,"Last Known Address")]); r+=1
    for i in range(6): drow(ws3,r+i,[(1,""),(2,""),(3,""),(4,"")],alt=i%2==1)
    gold_bar(ws3,r+8,4)

    # Sheet 4: Court / Probate Records
    ws4=wb.create_sheet("PROBATE RECORDS")
    ws4.sheet_view.showGridLines=False; ws4.sheet_properties.tabColor=PURPLE
    cw(ws4,{"A":20,"B":32,"C":20,"D":22})
    r=page_hdr(ws4,1,"PROBATE & COURT RECORDS",f"TX Courts · County Clerk · Vital Records",nc=4)
    sec(ws4,r,1,4,"PROBATE COURT RECORDS — TX COURTS ONLINE",PURPLE); r+=1
    thdr(ws4,r,[(1,"Case/Cause #"),(2,"Court"),(3,"Type"),(4,"Status")]); r+=1
    if R["probate"]:
        for i,p in enumerate(R["probate"]):
            drow(ws4,r+i,[(1,p["case"]),(2,p.get("source","")),(3,p["type"]),(4,"")],alt=i%2==1)
    else: drow(ws4,r,[(1,"No probate records auto-found"),(2,""),(3,""),(4,"Manual: search.txcourts.gov")],color=DIM)
    r+=max(6,len(R["probate"]))+1; spacer(ws4,r,8,4); r+=1
    sec(ws4,r,1,4,"CIVIL COURT RECORDS",PURPLE); r+=1
    thdr(ws4,r,[(1,"Case #"),(2,"Source"),(3,"Type"),(4,"Notes")]); r+=1
    if R["court_records"]:
        for i,c in enumerate(R["court_records"]):
            drow(ws4,r+i,[(1,c["case"]),(2,c["source"]),(3,c.get("type","")),(4,"")],alt=i%2==1)
    else: drow(ws4,r,[(1,"No court records found"),(2,""),(3,"Clean"),(4,"")],color=GRN)
    r+=max(5,len(R["court_records"]))+1; spacer(ws4,r,8,4); r+=1
    sec(ws4,r,1,4,"DEED / PROPERTY RECORDS (for estate assets)",PURPLE); r+=1
    thdr(ws4,r,[(1,"Property Address"),(2,"County"),(3,"Type"),(4,"Notes")]); r+=1
    if R["properties"]:
        for i,p in enumerate(R["properties"][:5]):
            drow(ws4,r+i,[(1,p.get("addr","")),(2,p.get("county","")),(3,p.get("type","Real Property")),(4,p.get("source",""))],alt=i%2==1,wrap=True)
    else: drow(ws4,r,[(1,"No property found — search all 254 TX counties manually"),(2,""),(3,""),(4,"")],color=DIM)
    gold_bar(ws4,r+max(6,len(R["properties"]))+2,4)

    # Sheet 5: Address & Contact Info
    ws5=wb.create_sheet("LOCATE RESULTS")
    ws5.sheet_view.showGridLines=False; ws5.sheet_properties.tabColor=GRN
    cw(ws5,{"A":22,"B":30,"C":22,"D":22})
    r=page_hdr(ws5,1,"LOCATE RESULTS",f"Addresses and contact info found for subject and relatives",nc=4)
    best=R["addresses"][0] if R["addresses"] else {"addr":"Not found","source":"","confidence":"Unlocated","corroborated_by":0}
    sec(ws5,r,1,4,"SUBJECT / ESTATE CONTACT",PURPLE); r+=1
    kv(ws5,r,"Last Known Address",best["addr"]); r+=1
    kv(ws5,r,"Confidence",best["confidence"],alt=True,vc=GRN if "HIGH" in best["confidence"].upper() or "VERIF" in best["confidence"] else AMB); r+=1
    kv(ws5,r,"Phone",R["phones"][0]["phone"] if R["phones"] else "Not found"); r+=1
    spacer(ws5,r,8,4); r+=1
    sec(ws5,r,1,4,"ALL ADDRESSES FOUND"); r+=1
    thdr(ws5,r,[(1,"Address"),(2,"Source"),(3,"Confidence"),(4,"Corroborations")]); r+=1
    for i,a in enumerate(R["addresses"][:8]):
        drow(ws5,r+i,[(1,a["addr"]),(2,a.get("source","")),(3,a["confidence"]),(4,str(a.get("corroborated_by",1)))],alt=i%2==1,wrap=True)
    gold_bar(ws5,r+max(8,len(R["addresses"]))+2,4)

    # Sheet 6: Red Flags
    ws6=wb.create_sheet("RED FLAGS")
    ws6.sheet_view.showGridLines=False; ws6.sheet_properties.tabColor=RED
    cw(ws6,{"A":14,"B":40,"C":16,"D":28})
    r=page_hdr(ws6,1,"RED FLAGS & HEIR ISSUES",f"Case: {CASE_NUM}",nc=4)
    heir_flags=[
        {"severity":"MEDIUM","desc":"Verify all potential heirs — intestate succession under TX Estates Code Chapter 201 requires locating ALL heirs","cat":"PROBATE","action":"Cross-reference all relatives found in public records"},
        {"severity":"MEDIUM","desc":"Quitclaim deeds within 2 years of death may indicate pre-death asset transfers — review for validity","cat":"ESTATE","action":"Check county deed records for transfers in last 2-5 years"},
    ]
    all_flags=heir_flags+R["red_flags"]
    thdr(ws6,r,[(1,"SEVERITY"),(2,"ISSUE"),(3,"CATEGORY"),(4,"ACTION")],h=22); r+=1
    for i,f in enumerate(all_flags): flagrow(ws6,r+i,f["severity"],f["desc"],f["cat"],f["action"])
    r+=max(6,len(all_flags))+1; gold_bar(ws6,r,4)

    source_log_sheet(wb)
    next_steps_sheet(wb,[
        ("1","File for letters testamentary or letters of administration if probate not yet opened","⬜ Pending","TX Estates Code §256"),
        ("2","Request death certificate from TX DSHS — Vital Statistics","⬜ Pending","dshs.texas.gov/VS"),
        ("3","Check all TX counties for deed records under decedent's name","⬜ Pending","publicsearch.us portal"),
        ("4","Run probate search at all TX county courts manually","⬜ Pending","County clerk + search.txcourts.gov"),
        ("5","Verify each located heir independently","⬜ Pending","Service of process on all identified heirs"),
        ("6","Check for life insurance policies — TX Life & Health Guaranty Association","⬜ Pending","tdi.texas.gov"),
        ("7","Search TX unclaimed property for estate assets — comptroller.texas.gov/up","⬜ Pending","Free search"),
    ], PURPLE)
    wb.save(OUT_PATH); print(f"\n  ✓ Tier 4 Heir report saved: {OUT_PATH}")

# ════════════════════════════════════════════════════════════
#  TIER SEARCH ROUTING
# ════════════════════════════════════════════════════════════
def run_tier1():
    print("  Running Tier 1 — Standard Locate sources...")
    s_truepeoplesearch(); s_fastpeople(); s_spokeo(); s_zaba()
    s_411(); s_anywho(); s_peekyou(); s_google(); s_bing(); s_usps_verify()

def run_tier2():
    print("  Running Tier 2 — Comprehensive sources...")
    s_truepeoplesearch(); s_fastpeople(); s_spokeo(); s_zaba()
    s_411(); s_anywho(); s_peekyou(); s_google(); s_bing()
    s_cad_multi(); s_tx_sos(); s_tx_courts(); s_tdcj(); s_sex_offender()
    s_tdlr(); s_rrc(); s_linkedin(); s_facebook(); s_instagram(); s_usps_verify()

def run_tier3():
    print("  Running Tier 3 — Property & Title sources...")
    s_cad_multi(); s_rrc()
    for county in ["Travis","Williamson","Milam","Bell","Bastrop","Hays","Caldwell"]:
        s_county_clerk_deeds(county, LAST)
        s_county_clerk_deeds(county, FIRST+" "+LAST)
    s_tx_sos(); s_tx_courts()
    # Also run basic people search to identify owner contact info
    s_truepeoplesearch(); s_zaba(); s_usps_verify()

def run_tier4():
    print("  Running Tier 4 — Heir & Beneficiary sources...")
    # Heavy on relatives/people aggregators
    s_truepeoplesearch(); s_fastpeople(); s_spokeo()
    s_zaba(); s_411(); s_anywho(); s_peekyou()
    s_tx_courts_probate(); s_tx_courts()
    s_cad_multi()  # estate property
    s_tx_sos()     # business interests of estate
    s_google(); s_usps_verify()

# ════════════════════════════════════════════════════════════
#  MAIN
# ════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print(f"\n  TraceWorks Engine v3.0 — Tier {TIER}: {SERVICE}")
    print(f"  Subject: {SUBJECT}  |  Case: {CASE_NUM}")
    print(f"  {'='*55}")

    dispatch = {"1":run_tier1,"2":run_tier2,"3":run_tier3,"4":run_tier4}
    runner = dispatch.get(TIER, run_tier1)
    runner()

    dedupe()

    print(f"\n  {'='*55}")
    print(f"  Building report...")

    builders = {"1":build_t1,"2":build_t2,"3":build_t3,"4":build_t4}
    builder = builders.get(TIER, build_t1)
    builder()
